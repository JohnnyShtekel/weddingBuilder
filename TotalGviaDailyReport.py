# coding=utf-8
import pandas as pd
from esg_services import EsgServiceManager
import numpy as np
from StyleFrame import StyleFrame
from pandas import concat
from openpyxl import load_workbook
from openpyxl.styles import Protection
import datetime
import os


class TotalGviaDailyReport(object):
    def __init__(self, chosen_date, for_previous_month):
        self.for_previous_month = for_previous_month
        self.chosen_date = chosen_date
        self.year = str(chosen_date.year)
        self.month = '0' + str(chosen_date.month) if chosen_date.month < 10 else str(chosen_date.month)
        self.day = '0' + str(chosen_date.day) if chosen_date.day < 10 else str(chosen_date.day)
        self.manager = EsgServiceManager()
        self.df = pd.DataFrame()
        self.rows_of_sum = []
        self.df_from_sql_for_gvia_megvia_and_tzefi = np.nan
        query_for_vat = "SELECT * from dbo.TaxRate"
        q = self.manager.db_service.search(query=query_for_vat)
        df_for_vat = pd.DataFrame.from_records(q)
        self.vat = df_for_vat['Tax'][len(df_for_vat) - 1]

    def add_col_team(self):
        query = '''SELECT NameTeam [צוות]
            FROM dbo.tblTeamList
            WHERE NameTeam  NOT LIKE 'כללי'
            ORDER BY NameTeam'''
        q = self.manager.db_service.search(query=query)
        self.df = pd.DataFrame.from_records(q)

    def get_paid_payment_for_consultation_and_gvia_from_excel_daily_report(self):
        df = pd.read_excel(u' דוח גביה יומי {day}-{month}-{year}.xlsx'.format(day=self.chosen_date.day, month=self.chosen_date.month,
                                                                              year=self.chosen_date.year), convert_float=True,
                           sheetname=u'דוח גביה יומי חיובי')
        df = df[[u'צוות', u'שם לקוח', u'סוג לקוח', u'תשלום ששולם עד היום לייעוץ', u'תשלום ששולם עד היום לגביה']]
        list_of_sum_rows = []
        for row in range(0, len(df)):
            if df[u'שם לקוח'][row].startswith(u'סיכום'):
                list_of_sum_rows.append(row)
        df.drop(df.index[list_of_sum_rows], inplace=True)
        df.index = range(len(df))
        return df

    def add_gvia_cols_for_3_kinds_of_customers(self):
        df = self.get_paid_payment_for_consultation_and_gvia_from_excel_daily_report()
        list_for_regular_customer_col = []
        list_for_consultation_customer_col = []
        list_for_project_customer_col = []
        team = df[u'צוות'][0]
        regular_sum = 0
        consultation_sum = 0
        project_sum = 0
        for row in range(0, len(df)):
            if df[u'צוות'][row] == team:
                if str(df[u'תשלום ששולם עד היום לייעוץ'][row]) == "nan":
                    df[u'תשלום ששולם עד היום לייעוץ'][row] = 0.0
                if df[u'סוג לקוח'][row] == u'ES - בסיס הצלחה':
                    regular_sum += df[u'תשלום ששולם עד היום לייעוץ'][row]
                elif df[u'סוג לקוח'][row] == u'ES-יעוץ':
                    consultation_sum += df[u'תשלום ששולם עד היום לייעוץ'][row]
                elif df[u'סוג לקוח'][row] == u'פרוייקטלי':
                    project_sum += df[u'תשלום ששולם עד היום לייעוץ'][row]
            if row == len(df) - 1:
                list_for_regular_customer_col.append(regular_sum)
                list_for_consultation_customer_col.append(consultation_sum)
                list_for_project_customer_col.append(project_sum)
            elif df[u'צוות'][row + 1] != team:
                team = df[u'צוות'][row + 1]
                list_for_regular_customer_col.append(regular_sum)
                list_for_consultation_customer_col.append(consultation_sum)
                list_for_project_customer_col.append(project_sum)
                regular_sum = 0
                consultation_sum = 0
                project_sum = 0
        self.df[u'גביה מרגיל'] = list_for_regular_customer_col
        self.df[u'גביה מייעוץ'] = list_for_consultation_customer_col
        self.df[u'גביה מפרויקטלי'] = list_for_project_customer_col

    def add_col_total_gvia(self):
        list_of_formulas = []
        for row in range(0, len(self.df)):
            formula = "=B{row}+C{row}+D{row}".format(row=row + 2)
            list_of_formulas.append(formula)
        self.df[u'גביה כוללת'] = list_of_formulas

    def get_df_from_sql_for_gvia_megvia_and_tzefi(self):
        # TODO: If you want to get last month report, change the month below:
        query = '''SELECT PaySuccessFU, PaySuccessFUTeam, DateFU, dbo.tblCustomers.KodCustomer, NameCustomer,NameTeam
          FROM dbo.tbltaxes
          INNER JOIN dbo.tblCustomers
            ON dbo.tbltaxes.KodCustomer = dbo.tblCustomers.KodCustomer
          INNER JOIN dbo.tblTeamList
            ON dbo.tblCustomers.KodTeamCare = dbo.tblTeamList.KodTeamCare
        INNER JOIN dbo.tblAgreementConditionAdvice
            ON dbo.tblAgreementConditionAdvice.KodCustomer = dbo.tblCustomers.KodCustomer
          WHERE YEAR(DateFU) = YEAR(CAST('{year}-{month}-{day} 00:00:00' as DATETIME)) AND MONTH(DateFU) = MONTH(CAST('{year}-{month}-{day} 00:00:00' as DATETIME))
        ORDER BY NameTeam'''.format(day=self.day, month=self.month, year=self.year)
        print query
        q = self.manager.db_service.search(query=query)
        self.df_from_sql_for_gvia_megvia_and_tzefi = pd.DataFrame.from_records(q)
        print self.df_from_sql_for_gvia_megvia_and_tzefi

    def add_col_gvia_from_gvia(self):
        df = (self.df_from_sql_for_gvia_megvia_and_tzefi.groupby('NameTeam').PaySuccessFU.sum()).reset_index()
        df['PaySuccessFU'] = df['PaySuccessFU'].apply(lambda x: x / (1 + self.vat))
        dict_for_consultation = df.set_index('NameTeam')['PaySuccessFU'].to_dict()
        self.df[u'גביה מהגביה'] = np.nan
        for row in range(0, len(self.df)):
            if self.df[u'צוות'][row] in dict_for_consultation.keys():
                self.df.set_value(row, u'גביה מהגביה',
                                  round(dict_for_consultation.get(self.df[u'צוות'][row])))


    def add_col_tzefi_from_gvia(self):
        df = (self.df_from_sql_for_gvia_megvia_and_tzefi.groupby('NameTeam').PaySuccessFUTeam.sum()).reset_index()
        df['PaySuccessFUTeam'] = df['PaySuccessFUTeam'].apply(lambda x: x / (1 + self.vat))
        dict_for_consultation = df.set_index('NameTeam')['PaySuccessFUTeam'].to_dict()
        self.df[u'צפי של הגביה / צוות'] = np.nan
        for row in range(0, len(self.df)):
            if self.df[u'צוות'][row] in dict_for_consultation.keys():
                self.df.set_value(row, u'צפי של הגביה / צוות',
                                  round(dict_for_consultation.get(self.df[u'צוות'][row])))

    def add_col_total_gvia_plus_tzefi_gvia_team(self):
        list_for_column = []
        for row in range(0, len(self.df)):
            list_for_column.append("=E{row}+G{row}".format(row=row + 2))
        self.df[u'גביה כוללת + צפי של הגבייה / צוות'] = list_for_column

    def add_col_target(self):
        list_of_targets = []
        list_of_indexes_for_teams_targets = [3, 2, 1, 10, 6, 4, 5, 7, 8]
        # TODO: If you want to get last month report, change the month below:
        query = '''SELECT *
        FROM dbo.tblMonthlyTarget
        WHERE year = YEAR(CAST('{year}-{month}-{day} 00:00:00' as DATETIME)) AND month = MONTH(CAST('{year}-{month}-{day} 00:00:00' as DATETIME))'''.format(day=self.day, month=self.month, year=self.year)
        q = self.manager.db_service.search(query=query)
        df = pd.DataFrame.from_records(q)
        for index in list_of_indexes_for_teams_targets:
            list_of_targets.append(df['sumOfTarget{index}'.format(index=index)][0])
        self.df[u'יעד'] = list_of_targets

    def add_col_percentage_of_completion(self):
        self.df[u'אחוז ביצוע'] = np.nan
        for row in range(0, len(self.df)):
            self.df[u'אחוז ביצוע'][row] = "=(IF(I{row}>0,E{row}/I{row},0))".format(row=row + 2)


    def add_col_percentage_of_completion_including_tzefi(self):
        self.df[u'אחוז ביצוע כולל צפי'] = np.nan
        for row in range(0, len(self.df)):
            self.df[u'אחוז ביצוע כולל צפי'][row] = "=(IF(I{row}>0,H{row}/I{row},0))".format(row=row + 2)

    def add_row_for_yesum(self):
        formula = "=SUM(({col}2:{col}5),({col}8,{col}10))"
        new_yesum_row = pd.DataFrame({u'צוות': [u'סה"כ יישום'], u'גביה מרגיל': [formula.format(col='B')],
                          u'גביה מייעוץ': [formula.format(col='C')], u'גביה מפרויקטלי': [formula.format(col='D')],
                          u'גביה כוללת': [formula.format(col='E')], u'גביה מהגביה': [formula.format(col='F')],
                          u'צפי של הגביה / צוות': [formula.format(col='G')],
                          u'גביה כוללת + צפי של הגבייה / צוות': [formula.format(col='H')],
                          u'יעד': [formula.format(col='I')], u'אחוז ביצוע': ["=(IF(I11>0,E11/I11,0))"],
                          u'אחוז ביצוע כולל צפי': ["=(IF(I11>0,H11/I11,0))"]})
        self.df = concat([self.df, new_yesum_row]).reset_index(drop=True)

    def add_row_for_gvia(self):
        df = self.get_paid_payment_for_consultation_and_gvia_from_excel_daily_report()
        regular_sum = 0
        consultation_sum = 0
        project_sum = 0
        for row in range(0, len(df)):
            if str(df[u'תשלום ששולם עד היום לגביה'][row]) == "nan":
                df[u'תשלום ששולם עד היום לגביה'][row] = 0.0
            if df[u'סוג לקוח'][row] == u'ES - בסיס הצלחה':
                regular_sum += df[u'תשלום ששולם עד היום לגביה'][row]
            elif df[u'סוג לקוח'][row] == u'ES-יעוץ':
                consultation_sum += df[u'תשלום ששולם עד היום לגביה'][row]
            elif df[u'סוג לקוח'][row] == u'פרוייקטלי':
                project_sum += df[u'תשלום ששולם עד היום לגביה'][row]
        formula = "=SUM({col}2:{col}10)"
        new_gvia_row = pd.DataFrame({u'צוות': [u'סה"כ גביה'], u'גביה מרגיל': [regular_sum],
                                      u'גביה מייעוץ': [consultation_sum],
                                      u'גביה מפרויקטלי': [project_sum],
                                      u'גביה כוללת': ["=B12+C12+D12"],
                                      u'גביה מהגביה': [formula.format(col='F')],
                                      u'צפי של הגביה / צוות': [formula.format(col='G')],
                                      u'גביה כוללת + צפי של הגבייה / צוות': [formula.format(col='H')],
                                      u'יעד': [formula.format(col='I')], u'אחוז ביצוע': ["=(IF(I12>0,E12/I12,0))"],
                                      u'אחוז ביצוע כולל צפי': ["=(IF(I12>0,H12/I12,0))"]})
        self.df = concat([self.df, new_gvia_row]).reset_index(drop=True)



    def order_columns(self):
        self.df = self.df[[u'צוות', u'גביה מרגיל', u'גביה מייעוץ', u'גביה מפרויקטלי', u'גביה כוללת', u'גביה מהגביה',
                           u'צפי של הגביה / צוות', u'גביה כוללת + צפי של הגבייה / צוות', u'יעד', u'אחוז ביצוע',
                           u'אחוז ביצוע כולל צפי']]

    def change_types_of_cells(self):
        wb = load_workbook(u'דוח גביה יומי מחלקתי {day}-{month}-{year}.xlsx'.format(day=self.chosen_date.day, month=self.chosen_date.month,
                                                                                  year=self.chosen_date.year))
        ws = wb.active
        for i in range(2, 13):
            ws['J{i}'.format(i=i)].number_format = '0%'
            ws['K{i}'.format(i=i)].number_format = '0%'
            ws['B{i}'.format(i=i)].number_format = '#,###'
            ws['C{i}'.format(i=i)].number_format = '#,###'
            ws['D{i}'.format(i=i)].number_format = '#,###'
            ws['E{i}'.format(i=i)].number_format = '#,###'
            ws['F{i}'.format(i=i)].number_format = '#,###'
            ws['G{i}'.format(i=i)].number_format = '#,###'
            ws['H{i}'.format(i=i)].number_format = '#,###'
            ws['I{i}'.format(i=i)].number_format = '#,###'
        for column in ws.columns:
            for cell in column:
                if not cell.protection.locked:
                    cell.protection = Protection(locked=True)
        wb.save(u'דוח גביה יומי מחלקתי {day}-{month}-{year}.xlsx'.format(day=self.chosen_date.day, month=self.chosen_date.month,
                                                                                  year=self.chosen_date.year))

    def create_df_for_mail_team(self, team):
        team_df = self.df[self.df[u'צוות'] == team].reset_index(drop=True)
        team_df[u'גביה כוללת'][0] = format(int(round(team_df[u'גביה מרגיל'][0] + team_df[u'גביה מייעוץ'][0] +
                                                     team_df[u'גביה מפרויקטלי'][0])), ',d')
        team_df[u'גביה מייעוץ'][0] = format(int(round(team_df[u'גביה מייעוץ'][0])), ',d')
        team_df[u'גביה מפרויקטלי'][0] = format(int(round(team_df[u'גביה מפרויקטלי'][0])), ',d')
        team_df[u'גביה מהגביה'][0] = format(int(round(team_df[u'גביה מהגביה'][0])), ',d')
        team_df[u'צפי של הגביה / צוות'][0] = format(int(round(team_df[u'צפי של הגביה / צוות'][0])), ',d')
        team_df[u'גביה כוללת + צפי של הגבייה / צוות'][0] = format(int(team_df[u'גביה כוללת'][0].
                                                                      replace(',', '')) +
                                                                  int(team_df[u'צפי של הגביה / צוות'][0].replace
                                                                      (',', '')), ',d')
        team_df[u'יעד'][0] = format(int(round(team_df[u'יעד'][0])), ',d')
        if int(team_df[u'יעד'][0].replace(',', '')) > 0:
            percentage_of_completion = round(
                100 * (float(team_df[u'גביה כוללת'][0].replace(',', '')) / float(team_df[u'יעד'][0].replace(',', ''))))
            str_for_percentage_of_completion = ''.join([str(percentage_of_completion)[:-2], "%"])
            team_df[u'אחוז ביצוע'][0] = str_for_percentage_of_completion
        else:
            team_df[u'אחוז ביצוע'][0] = 0
        if int(team_df[u'יעד'][0].replace(',', '')) > 0:
            percentage_of_completion_including_tzefi = round(100 *
                                                             float(team_df[u'גביה כוללת + צפי של הגבייה / צוות'][
                                                                       0].replace(',', '')) /
                                                             float(team_df[u'יעד'][0].replace(',', '')))
            str_for_percentage_of_completion_including_tzefi = ''.join(
                [str(percentage_of_completion_including_tzefi)[:-2], "%"])
            team_df[u'אחוז ביצוע כולל צפי'][0] = str_for_percentage_of_completion_including_tzefi
        else:
            team_df[u'אחוז ביצוע כולל צפי'][0] = 0
        team_df.set_index(u'צוות', inplace=True)
        team_df = team_df.rename_axis(None)
        return team_df

    def send_mail_for_each_team(self):
        teams_to_emails = {u'אודם': ['Ilyag@esg.co.il'], u'אודם+שנהב': ['roeeE@esg.co.il'], u'ברקת': ['yanivm@esg.co.il'],
                           u'אלמוג': ['ZviL@esg.co.i'], u'קריסטל': ['davidme@esg.co.il'],
                           u'טורקיז': ['Gils@esg.co.il'], u'טורקיז+ברקת': ['michalz@esg.co.il'], u'שנהב': ['elinorn@esg.co.il'],
                           u'שוהם - שכר': ['yaelz@bsachar.co.il']}
        df = pd.read_excel(u'דוח גביה יומי מחלקתי {day}-{month}-{year}.xlsx'.format(day=self.chosen_date.day, month=self.chosen_date.month,
                                                                         year=self.chosen_date.year), convert_float=True,
                      sheetname=u'דוח גביה יומי לפי צוותים')
        for key in teams_to_emails:
            files = []
            if '+' in key:
                team1 = key[:key.index('+')]
                team2 = key[key.index('+') + 1:]
                team_df1 = self.create_df_for_mail_team(team1)
                team_df2 = self.create_df_for_mail_team(team2)
                df_for_2_teams = concat([team_df1, team_df2])
                df_for_2_teams.drop(df_for_2_teams.columns[[0, 1, 2]], axis=1, inplace=True)
                template = df_for_2_teams.to_html()
                style = u'''
                    <style>
                        table, th, td {
                            border: 1px solid black;
                            text-align: center;
                            align: left;
                        }
                    </style>
                '''
                template = style + u'\n' + template
                file_for_team1 = u' דוח גביה יומי {day}-{month}-{year} צוות {team}.xlsx'.format(team=team1, day=self.chosen_date.day,
                                                                                           month=self.chosen_date.month,
                                                                                           year=self.chosen_date.year)
                file_for_team2 = u' דוח גביה יומי {day}-{month}-{year} צוות {team}.xlsx'.format(team=team2, day=self.chosen_date.day,
                                                                                           month=self.chosen_date.month,
                                                                                           year=self.chosen_date.year)
                files.append(file_for_team1)
                files.append(file_for_team2)
            else:
                team_df = self.create_df_for_mail_team(key)
                team_df.drop(team_df.columns[[0, 1, 2]], axis=1, inplace=True)
                template = team_df.to_html()
                style = u'''
                    <style>
                        table, th, td {
                            border: 1px solid black;
                            text-align: center;
                            align: left;
                        }
                    </style>
                '''
                template = style + u'\n' + template
                file_name = u' דוח גביה יומי {day}-{month}-{year} צוות {team}.xlsx'.format(team=key, day=self.chosen_date.day,
                                                                                           month=self.chosen_date.month,
                                                                                           year=self.chosen_date.year)
                files.append(file_name)
            # TODO: change mail_to to emails list of the current key!!!
            mail_to = ['yakin@esg.co.il']
            mail_from = ('ESG Server', 'server@esg.co.il')
            subject = u' דוח גביה יומי {day}-{month}-{year}'.format(day=self.chosen_date.day, month=self.chosen_date.month, year=self.chosen_date.year)
            html = template
            response = self.manager.mail_service.send_mail(mail_from=mail_from, mail_to=mail_to, files=files,
                                                          subject=subject, html=html)
            print response.content

    def send_mail_to_managers(self):
        files = []
        emails = ['roeeE@esg.co.il', 'michalz@esg.co.il', 'avi@esg.co.il', 'NatanF@esg.co.il', 'YakiN@esg.co.il',
                  'rond@esg.co.il', 'HaniP@esg.co.il']
        file_name = u'דוח גביה יומי מחלקתי {day}-{month}-{year}.xlsx'.format(day=self.chosen_date.day, month=self.chosen_date.month, year=self.chosen_date.year)
        files.append(file_name)
        # TODO: change mail_to to emails list!!!
        mail_to = ['yakin@esg.co.il']
        mail_from = ('ESG Server', 'server@esg.co.il')
        subject = u' דוח גביה יומי לפי מחלקות {day}-{month}-{year}'.format(day=self.chosen_date.day, month=self.chosen_date.month, year=self.chosen_date.year)
        response = self.manager.mail_service.send_mail(mail_from=mail_from, mail_to=mail_to, files=files,
                                                           subject=subject, text=u'מצורף דוח גביה יומי לפי מחלקות')
        print response.content



    def create_final_excel_file(self):
        sf = StyleFrame(self.df)
        reportName = u'דוח גביה יומי מחלקתי {day}-{month}-{year}.xlsx'.format(day=self.chosen_date.day,
                                                                                                 month=self.chosen_date.month,
                                                                                                 year=self.chosen_date.year)
        writer = StyleFrame.ExcelWriter(reportName)
        sf.to_excel(excel_writer=writer, sheet_name=u'דוח גביה יומי לפי צוותים', right_to_left=True,
                    row_to_add_filters=0,
                    columns_and_rows_to_freeze='A2')
        writer.save()
        return reportName


    def run_gvia_total_report(self):
        self.add_col_team()
        self.add_gvia_cols_for_3_kinds_of_customers()
        self.add_col_total_gvia()
        self.get_df_from_sql_for_gvia_megvia_and_tzefi()
        self.add_col_gvia_from_gvia()
        self.add_col_tzefi_from_gvia()
        self.add_col_total_gvia_plus_tzefi_gvia_team()
        self.add_col_target()
        self.add_col_percentage_of_completion()
        self.add_col_percentage_of_completion_including_tzefi()
        self.add_row_for_yesum()
        self.add_row_for_gvia()
        self.order_columns()
        reportName = self.create_final_excel_file()
        self.change_types_of_cells()
        if not self.for_previous_month:
            self.send_mail_for_each_team()
            self.send_mail_to_managers()
        return reportName




if __name__ == '__main__':

    chosen_date = datetime.datetime(2016,8,25)
    gvia_teams = TotalGviaDailyReport(chosen_date, False)
    gvia_teams.add_col_team()
    gvia_teams.add_gvia_cols_for_3_kinds_of_customers()
    gvia_teams.add_col_total_gvia()
    gvia_teams.get_df_from_sql_for_gvia_megvia_and_tzefi()
    gvia_teams.add_col_gvia_from_gvia()
    gvia_teams.add_col_tzefi_from_gvia()
    gvia_teams.add_col_total_gvia_plus_tzefi_gvia_team()
    gvia_teams.add_col_target()
    gvia_teams.add_col_percentage_of_completion()
    gvia_teams.add_col_percentage_of_completion_including_tzefi()
    gvia_teams.add_row_for_yesum()
    gvia_teams.add_row_for_gvia()
    gvia_teams.order_columns()

    sf = StyleFrame(gvia_teams.df)
    writer = StyleFrame.ExcelWriter(u'דוח גביה יומי מחלקתי {day}-{month}-{year}.xlsx'.format(day=gvia_teams.chosen_date.day,
                                                                                       month=gvia_teams.chosen_date.month,
                                                                                       year=gvia_teams.chosen_date.year))
    sf.to_excel(excel_writer=writer, sheet_name=u'דוח גביה יומי לפי צוותים', right_to_left=True, row_to_add_filters=0,
            columns_and_rows_to_freeze='A2')
    writer.save()
    gvia_teams.change_types_of_cells()
    # gvia_teams.send_mail_for_each_team()
    # gvia_teams.send_mail_to_managers()








