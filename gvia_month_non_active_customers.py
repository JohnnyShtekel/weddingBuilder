# coding=utf-8
import pandas as pd
from esg_services import EsgServiceManager
from StyleFrame import StyleFrame, colors
from pandas import concat
import re
from openpyxl import load_workbook
from openpyxl.styles import Protection
import datetime



class GviaMonthNonActiveReport(object):
    def __init__(self, query):
        self.manager = EsgServiceManager()
        self.df = pd.DataFrame.from_records(self.manager.db_service.search(query=query))
        self.rows_of_sum = []
        query_for_vat = "SELECT * from dbo.TaxRate"
        q = self.manager.db_service.search(query=query_for_vat)
        df_for_vat = pd.DataFrame.from_records(q)
        self.vat = df_for_vat['Tax'][len(df_for_vat) - 1]

    def subtract_vat(self):
        self.df[u'תשלום על בונוס'] = self.df[u'תשלום על בונוס'].apply(lambda x: x / (1 + self.vat))
        self.df[u'סכום התשלומים מעבר לאשראי'] = self.df[u'סכום התשלומים מעבר לאשראי'].apply(lambda x: x / (1 + self.vat))
        self.df[u'תשלומים מיוחדים'] = self.df[u'תשלומים מיוחדים'].apply(lambda x: x / (1 + self.vat))

    def add_col_tzefi_for_month(self, status):
        tzefi_for_month = []
        rows = len(self.df.index)
        if status == 1:
            for r in range(2, rows + 2):
                if r-2 not in self.rows_of_sum:
                    tzefi_for_month.append('=IF(E{r}={yes},0,IF(I{r}>3,0,F{r}))'.format(r=r, yes='"כן"'))
                else:
                    tzefi_for_month.append(self.df.iloc[r - 2][u'צפי לחודש'])
        else:
            for r in range(2, rows + 2):
                tzefi_for_month.append(0)
        self.df[u'צפי לחודש'] = tzefi_for_month


    def add_col_additional_payments_for_month(self, status):
        additional_payments = []
        rows = len(self.df.index)
        if status == 1:
            for r in range(2, rows + 2):
                if r-2 not in self.rows_of_sum:
                    additional_payments.append("=IF(E{r}={yes}, 0, SUM(G{r}+H{r}+J{r}))".format(r=r, yes='"כן"'))
                else:
                    additional_payments.append(self.df.iloc[r - 2][u'תשלומים נוספים בחודש'])
        else:
            for r in range(2, rows + 2):
                additional_payments.append(0)
        self.df[u'תשלומים נוספים בחודש'] = additional_payments

    def add_col_tzefi_meuhad_for_month(self, status):
        tzefi_meuhad = []
        rows = len(self.df.index)
        if status == 1:
            for r in range(2, rows + 2):
                if r-2 not in self.rows_of_sum:
                    tzefi_meuhad.append('=SUM(M{r}+L{r})'.format(r=r))
                else:
                    tzefi_meuhad.append(self.df.iloc[r - 2][u'צפי מאוחד'])
        else:
            for r in range(2, rows + 2):
                tzefi_meuhad.append(0)
        self.df[u'צפי מאוחד'] = tzefi_meuhad


    def set_3_last_notes(self):
        rows = len(self.df.index)
        for row in range(0, rows):
            last_3_notes = u""
            notes = self.df[u'הערות מגיליון הגביה'][row]
            patern = re.compile(r'\d{2}[/-]\d{2}[/-]\d{2,4}')
            list_of_notes = patern.split(notes)
            list_of_dates = re.findall(r'\d{2}[/-]\d{2}[/-]\d{2,4}', notes)
            list_of_notes = filter(lambda x: x != u'', list_of_notes[-3:])
            list_of_dates = filter(lambda x: x != u'', list_of_dates[-3:])
            for note, date in zip(list_of_notes, list_of_dates):
                last_3_notes = last_3_notes + date + note
            self.df[u'הערות מגיליון הגביה'][row] = last_3_notes


    def arrange_col_order(self):
        self.df = self.df[[u'צוות', u'שם לקוח', u'אמצעי תשלום', u'סוג לקוח', u'לקוח משפטי', u'תשלום ייעוץ חודשי',
                           u'תשלום על בונוס', u'תשלומים מיוחדים', u'מספר התשלומים מעבר לאשראי',
                           u'סכום התשלומים מעבר לאשראי', u'הערות', u'צפי לחודש', u'תשלומים נוספים בחודש', u'צפי מאוחד',
                           u'הערות מגיליון הגביה']]

    def change_type_of_col_to_int(self, col_name):
        self.df[col_name] = self.df[col_name].apply(lambda x: int(x))

    def change_type_of_sums_col_to_int(self):
        self.change_type_of_col_to_int(u'תשלום ייעוץ חודשי')
        self.change_type_of_col_to_int(u'תשלום על בונוס')
        self.change_type_of_col_to_int(u'תשלומים מיוחדים')
        self.change_type_of_col_to_int(u'סכום התשלומים מעבר לאשראי')



    def add_mid_sums(self):
        rows = len(self.df.index)
        team = self.df[u'צוות'][0]
        first_index = 0
        last_index = 0
        for r in range(0, rows):
            if r == rows - 1:
                last_index = len(self.df.index) - 1
                self.calc_mid_sums(first_index, last_index, team)
            if self.df[u'צוות'][r + 1] == team:
                last_index = last_index + 1

            else:
                if self.df[u'צוות'][r + 1] != team:
                    new_team = self.df[u'צוות'][r + 1]
                    self.calc_mid_sums(first_index, last_index, team)
                    last_index = last_index + 1
                    first_index = last_index + 1
                    team = new_team
        self.df.sort_index()

    def calc_mid_sums(self, first_index, last_index, team):
        index1 = first_index + 2
        index2 = last_index + 2
        sum_of_month_consultation = '=SUM(F{first_index}:F{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_pay_for_bonus = '=SUM(G{first_index}:G{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_special_payments = '=SUM(H{first_index}:H{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_month_over_credit = '=SUM(J{first_index}:J{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_month_zefi = '=SUM(L{first_index}:L{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_extra_payments_for_month = '=SUM(M{first_index}:M{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_zefi_meuhad = '=SUM(N{first_index}:N{last_index})'.format(first_index=index1, last_index=index2)
        new_sum_line = pd.DataFrame({u'צוות': [team], u'שם לקוח': [u'סיכום לצוות {team}'.format(team=team)],
                                     u'אמצעי תשלום': [""], u'סוג לקוח': [""], u'לקוח משפטי': [""],
                                     u'תשלום ייעוץ חודשי': [sum_of_month_consultation],
                                     u'תשלום על בונוס': [sum_of_pay_for_bonus],
                                     u'תשלומים מיוחדים': [sum_of_special_payments],
                                     u'מספר התשלומים מעבר לאשראי': [""],
                                     u'סכום התשלומים מעבר לאשראי': [sum_of_month_over_credit],
                                     u'הערות': [""], u'צפי לחודש': [sum_of_month_zefi],
                                     u'תשלומים נוספים בחודש': [sum_of_extra_payments_for_month],
                                     u'צפי מאוחד': [sum_of_zefi_meuhad], u'הערות מגיליון הגביה': [""]})
        new_df = concat([self.df.ix[:last_index], new_sum_line, self.df.ix[last_index + 1:]]).reset_index(drop=True)
        new_df.sort_index()
        self.df = new_df
        self.rows_of_sum.append(last_index + 1)

    def apply_style_on_sum_rows(self, sf):
        for row in self.rows_of_sum:
            sf.apply_style_by_indexes(indexes_to_style=[sf.index[row]], bg_color=colors.yellow, bold=True)

    def add_total_row(self):
        sum_of_month_consultation = '='
        sum_of_pay_for_bonus = '='
        sum_of_special_payments = '='
        sum_of_month_over_credit = '='
        sum_of_month_zefi = '='
        sum_of_extra_payments_for_month = '='
        sum_of_zefi_meuhad = '='
        for row in range(0, len(self.rows_of_sum) - 1):
            sum_of_month_consultation += "F" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_pay_for_bonus += "G" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_special_payments += "H" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_month_over_credit += "J" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_month_zefi += "L" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_extra_payments_for_month += "M" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_zefi_meuhad += "N" + str(self.rows_of_sum[row] + 2) + "+"
        sum_of_month_consultation = sum_of_month_consultation[:-1]
        sum_of_pay_for_bonus = sum_of_pay_for_bonus[:-1]
        sum_of_special_payments = sum_of_special_payments[:-1]
        sum_of_month_over_credit = sum_of_month_over_credit[:-1]
        sum_of_month_zefi = sum_of_month_zefi[:-1]
        sum_of_extra_payments_for_month = sum_of_extra_payments_for_month[:-1]
        sum_of_zefi_meuhad = sum_of_zefi_meuhad[:-1]
        new_sum_line = pd.DataFrame({u'צוות': [""], u'שם לקוח': [u'סיכום לכל הצוותים:'],
                                     u'אמצעי תשלום': [""], u'סוג לקוח': [""], u'לקוח משפטי': [""],
                                     u'תשלום ייעוץ חודשי': [sum_of_month_consultation],
                                     u'תשלום על בונוס': [sum_of_pay_for_bonus],
                                     u'תשלומים מיוחדים': [sum_of_special_payments],
                                     u'מספר התשלומים מעבר לאשראי': [""],
                                     u'סכום התשלומים מעבר לאשראי': [sum_of_month_over_credit],
                                     u'הערות': [""], u'צפי לחודש': [sum_of_month_zefi],
                                     u'תשלומים נוספים בחודש': [sum_of_extra_payments_for_month],
                                     u'צפי מאוחד': [sum_of_zefi_meuhad], u'הערות מגיליון הגביה': [""]})
        new_df = concat([self.df, new_sum_line]).reset_index(drop=True)
        self.df = new_df

    def get_df_for_bonus(self):
        query = '''SELECT
            NameCustomer,
            bonusFirstDAteTest AS [Bonus Start Date],
            ISNULL(bonusDAteTest, dateadd(MONTH, Amonth, CONVERT(DATE, DateNew))) AS [Bonus End Date],
            ISNULL(bonusTest/100.0,0) as [Bonus Test],
            DateP,
            DateInvoice,
            NumPay,
            Pay,
            SumS
        FROM
          dbo.tblCustomers
        INNER JOIN (
            SELECT *
            FROM
                (
                  SELECT
                    IDagreement,
                    KodCustomer,
                    DateNew,
                    bonusFirstDAteTest,
                    bonusDAteTest,
                    Amonth,
                    bonusTest,
                    agreementFinish,
                    row_number() OVER(PARTITION BY KodCustomer ORDER BY DateNew DESC) AS orderAgreementByDateDesc
                      FROM
                        dbo.tblAgreementConditionAdvice
                ) AS temp
            WHERE temp.orderAgreementByDateDesc = 1
            ) AS tblLastAgreements ON tblLastAgreements.KodCustomer = dbo.tblCustomers.KodCustomer
        INNER JOIN dbo.tblAgreementConditionAdvicePay
              ON tblLastAgreements.IDagreement = dbo.tblAgreementConditionAdvicePay.NumR
        WHERE CustomerStatus = 2
        AND closeCustomerDate IS NULL
        AND agreementFinish = 0
        ORDER BY NameCustomer, '''
        q = self.manager.db_service.search(query=query)
        df = pd.DataFrame.from_records(q)

    def get_bonus_date(self):
        last_agreement_bonus_begin_date = self.last_agreement.bonusFirstDAteTest
        last_agreement_bonus_end_date = self.last_agreement.bonusDAteTest
        if not last_agreement_bonus_begin_date:
            last_agreement_bonus_begin_date = self.last_agreement.DateNew
        if not last_agreement_bonus_end_date:
            last_agreement_bonus_end_date = self.last_agreement.DateNew + relativedelta(
                months=self.last_agreement.Amonth)
        return {'bonus_begin_date': last_agreement_bonus_begin_date,
                'bonus_end_date': last_agreement_bonus_end_date}

    def fulfilled_bonus_in_percent(self):
        tbl_agreement_condition_advice_pay = self.dal['tblAgreementConditionAdvicePay']
        if not self.last_agreement.bonusTest:
            return "אין בונוס"
        bonus_percent = self.last_agreement.bonusTest / 100.0
        invoices_rows = self.dal.session.query(tbl_agreement_condition_advice_pay).filter_by(
            NumR=self.last_agreement.IDagreement).all()
        # last_agreement = self.get_last_agreement_of_customer()
        last_agreement_bonus_date = self.get_bonus_date()
        last_agreement_bonus_begin_date = last_agreement_bonus_date['bonus_begin_date']
        last_agreement_bonus_end_date = last_agreement_bonus_date['bonus_end_date']
        sum_pay = sum(row.Pay for row in invoices_rows if row.Pay and row.NumPay is not None and row.NumPay != 0 and
                      last_agreement_bonus_begin_date <= row.DateP <= last_agreement_bonus_end_date)
        sum_efficiencies = sum(row.SumS for row in invoices_rows if
                               row.SumS is not None and row.DateInvoice is not None and last_agreement_bonus_begin_date <= datetime(
                                   year=row.DateInvoice.year, month=row.DateInvoice.month,
                                   day=1) <= last_agreement_bonus_end_date)
        bonus_test = sum_efficiencies - (sum_pay * bonus_percent)
        if bonus_test > 0 and self.last_agreement.bonus:
            if '%' in self.last_agreement.bonus:
                return bonus_test * float(self.last_agreement.bonus[:-1]) / 100.0
            else:
                return bonus_test * float(self.last_agreement.bonus) / 100.0
        else:
            return "לא עמד בדרישה"

    def change_types_of_cells(self):
        wb = load_workbook('gvia_month.xlsx')
        ws = wb.active
        for i in range(2, len(self.df) + 2):
            ws['F{i}'.format(i=i)].number_format = '#,###'
            ws['G{i}'.format(i=i)].number_format = '#,###'
            ws['H{i}'.format(i=i)].number_format = '#,###'
            ws['J{i}'.format(i=i)].number_format = '#,###'
            ws['L{i}'.format(i=i)].number_format = '#,###'
            ws['M{i}'.format(i=i)].number_format = '#,###'
            ws['N{i}'.format(i=i)].number_format = '#,###'
        for column in ws.columns:
            for cell in column:
                if not cell.protection.locked:
                    cell.protection = Protection(locked=True)
        wb.save('gvia_month.xlsx')

    def create_non_active_cutomers_gvia_month_report(self):
        self.subtract_vat()
        self.add_col_tzefi_for_month(0)
        self.add_col_additional_payments_for_month(0)
        self.add_col_tzefi_meuhad_for_month(0)
        ######################################################
        self.get_df_for_bonus()
        #######################################################
        self.set_3_last_notes()
        self.change_type_of_sums_col_to_int()
        self.change_type_of_col_to_int(u'מספר התשלומים מעבר לאשראי')
        self.add_mid_sums()
        self.rows_of_sum.append(len(self.df))
        self.add_total_row()
        self.add_col_tzefi_for_month(1)
        self.add_col_additional_payments_for_month(1)
        self.add_col_tzefi_meuhad_for_month(1)
        self.arrange_col_order()
        sf = StyleFrame(self.df)
        self.apply_style_on_sum_rows(sf)
        return sf














