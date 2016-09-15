# coding=utf-8
import pandas as pd
from esg_services import EsgServiceManager
from StyleFrame import StyleFrame, colors
from pandas import concat
import re
from openpyxl import load_workbook
from openpyxl.styles import Protection
import datetime
from gvia_month_non_active_customers import GviaMonthNonActiveReport
import numpy as np
from openpyxl.styles import Alignment



class GviaMonthReport(object):
    def __init__(self, chosen_date):
        self.chosen_date = chosen_date
        self.year = str(chosen_date.year)
        self.month = '0' + str(chosen_date.month) if chosen_date.month < 10 else str(chosen_date.month)
        self.day = '0' + str(chosen_date.day) if chosen_date.day < 10 else str(chosen_date.day)

        query = '''SELECT
  NameTeam AS [צוות],
  NameCustomer AS [שם לקוח],
  MiddlePay AS [אמצעי תשלום],
  AgreementKind AS [סוג לקוח],
  CASE WHEN proceFinished=0 THEN 'כן' ELSE '' END AS [לקוח משפטי],
  CASE WHEN Conditions='דולר' THEN ISNULL(PayDollar, 0)*4 ELSE ISNULL(PayDollar, 0) END AS [תשלום ייעוץ חודשי],
  ISNULL(SumBonus,0) as [תשלום על בונוס],
  ISNULL(SumSpecial,0) as [תשלומים מיוחדים],
  ISNULL(NumOfDebt,0) as [מספר התשלומים מעבר לאשראי],
  ISNULL(Debt,0) as [סכום התשלומים מעבר לאשראי],
  ISNULL(Remark,'') as [הערות],
  ISNULL(Text, '') AS [הערות מגיליון הגביה]
FROM
  dbo.tblCustomers
  LEFT JOIN
  dbo.tblTeamList
    ON dbo.tblCustomers.KodTeamCare = dbo.tblTeamList.KodTeamCare
  LEFT JOIN
  (SELECT * FROM
    (
      SELECT
        IDagreement,
        KodAgreementKind,
        KodCustomer,
        PayProcess,
        PayDollar,
        agreementFinish,
        row_number() OVER(PARTITION BY KodCustomer ORDER BY DateNew DESC) AS orderAgreementByDateDesc
      FROM
        dbo.tblAgreementConditionAdvice
    ) AS temp
      WHERE temp.orderAgreementByDateDesc = 1) AS tblLastAgreements
    ON dbo.tblCustomers.KodCustomer = tblLastAgreements.KodCustomer
  LEFT JOIN tblJudicialProc
    ON tblCustomers.KodCustomer = tblJudicialProc.kodCustomer
  LEFT JOIN dbo.tblMiddlePay
    ON tblLastAgreements.PayProcess = dbo.tblMiddlePay.KodMiddlePay
  LEFT JOIN dbo.tblAgreementKind
    ON dbo.tblAgreementKind.KodAgreementKind = tblLastAgreements.KodAgreementKind
  LEFT JOIN (
    SELECT DISTINCT NumR, Conditions,
      SUM(CASE WHEN NumPay = 0 AND SumP-Pay>0 AND PayRemark LIKE '%בונוס%' THEN SumP-pay ELSE 0 END) OVER(PARTITION BY NumR) AS SumBonus,
      SUM(CASE WHEN NumPay = 0 AND SumP-Pay>0  AND NOT PayRemark LIKE '%בונוס%' THEN SumP-pay ELSE 0 END) OVER(PARTITION BY NumR) AS SumSpecial,
      SUM(CASE WHEN NumPay > 0 AND SumP-Pay>0  AND DatePay IS NULL AND (CAST('{year}-{month}-{day} 00:00:00' AS DATETIME) > (CASE
     WHEN tblAgreementConditionAdvice.shotef = 1 THEN
       DateAdd(DAY, tblAgreementConditionAdvice.ashray+1,
               DateAdd(DAY,-1, Cast(CONVERT(VARCHAR(10),DATEPART(MM, DateAdd(month,1,tblAgreementConditionAdvicePay.DateP)),103)
                                    + '/' + '01/' + CONVERT(VARCHAR(10),DATEPART(YYYY, DateAdd(month,1,tblAgreementConditionAdvicePay.DateP)),103) AS DATETIME)))
     ELSE
       DateAdd(DAY, tblAgreementConditionAdvice.ashray+1,tblAgreementConditionAdvicePay.DateP)
     END)) THEN 1 ELSE 0 END) OVER(PARTITION BY NumR) AS NumOfDebt,
      SUM(CASE WHEN NumPay > 0 AND SumP-Pay>0  AND DatePay IS NULL AND (CAST('{year}-{month}-{day} 00:00:00' AS DATETIME) > (CASE
     WHEN tblAgreementConditionAdvice.shotef = 1 THEN
       DateAdd(DAY, tblAgreementConditionAdvice.ashray+1,
               DateAdd(DAY,-1, Cast(CONVERT(VARCHAR(10),DATEPART(MM, DateAdd(month,1,tblAgreementConditionAdvicePay.DateP)),103)
                                    + '/' + '01/' + CONVERT(VARCHAR(10),DATEPART(YYYY, DateAdd(month,1,tblAgreementConditionAdvicePay.DateP)),103) AS DATETIME)))
     ELSE
       DateAdd(DAY, tblAgreementConditionAdvice.ashray+1,tblAgreementConditionAdvicePay.DateP)
     END) ) THEN SumP ELSE 0 END) OVER(PARTITION BY NumR) AS Debt,
      STUFF((SELECT ', ' + tblRemarks.PayRemark AS [text()]FROM dbo.tblAgreementConditionAdvicePay tblRemarks WHERE tblRemarks.NumR = dbo.tblAgreementConditionAdvicePay.NumR AND tblRemarks.NumPay = 0 AND SumP-Pay>0 FOR XML PATH('')), 1, 1, '' ) AS Remark
    FROM dbo.tblAgreementConditionAdvicePay
    LEFT JOIN dbo.tblAgreementConditionAdvice
      ON dbo.tblAgreementConditionAdvice.IDagreement = dbo.tblAgreementConditionAdvicePay.NumR


     )AS tblSumAllPay
    ON tblLastAgreements.IDagreement = tblSumAllPay.NumR
  LEFT JOIN (
    SELECT KodCustomer, Text
    FROM tbltaxes
    ) tbltaxes
    ON
      dbo.tblCustomers.KodCustomer = tbltaxes.KodCustomer
WHERE CustomerStatus = 2
AND agreementFinish = 0
ORDER BY NameTeam'''.format(day=self.day, year=self.year, month=self.month)
        self.manager = EsgServiceManager()
        self.df = pd.DataFrame.from_records(self.manager.db_service.search(query=query))
        self.rows_of_sum = []
        query_for_vat = "SELECT * from dbo.TaxRate"
        q = self.manager.db_service.search(query=query_for_vat)
        df_for_vat = pd.DataFrame.from_records(q)
        self.vat = df_for_vat['Tax'][len(df_for_vat) - 1]

    def subtract_vat(self):
        self.df[u'תשלום על בונוס'] = self.df[u'תשלום על בונוס'].apply(lambda x: x / (1 + self.vat))
        self.df[u'סכום התשלומים מעבר לאשראי'] = self.df[u'סכום התשלומים מעבר לאשראי'].apply(
            lambda x: x / (1 + self.vat))
        self.df[u'תשלומים מיוחדים'] = self.df[u'תשלומים מיוחדים'].apply(lambda x: x / (1 + self.vat))

    def add_col_tzefi_for_month(self, status):
        tzefi_for_month = []
        rows = len(self.df.index)
        if status == 1:
            for r in range(2, rows + 2):
                if r - 2 not in self.rows_of_sum:
                    tzefi_for_month.append('=IF(E{r}={yes},0,IF(I{r}>3,0,F{r}))'.format(r=r, yes='"כן"'))
                else:
                    tzefi_for_month.append(self.df.iloc[r - 2][u'צפי לחודש'])
        else:
            for r in range(2, rows + 2):
                tzefi_for_month.append(0)
        self.df[u'צפי לחודש'] = tzefi_for_month

    def add_col_additional_payments_for_month(self, status):
        additional_payments = []
        rows = len(self.df.index)
        if status == 1:
            for r in range(2, rows + 2):
                if r - 2 not in self.rows_of_sum:
                    additional_payments.append("=IF(E{r}={yes}, 0, SUM(G{r}+H{r}+J{r}+M{r}))".format(r=r, yes='"כן"'))
                else:
                    additional_payments.append(self.df.iloc[r - 2][u'תשלומים נוספים בחודש'])
        else:
            for r in range(2, rows + 2):
                additional_payments.append(0)
        self.df[u'תשלומים נוספים בחודש'] = additional_payments

    def add_col_tzefi_meuhad_for_month(self, status):
        tzefi_meuhad = []
        rows = len(self.df.index)
        if status == 1:
            for r in range(2, rows + 2):
                if r - 2 not in self.rows_of_sum:
                    tzefi_meuhad.append('=SUM(N{r}+O{r})'.format(r=r))
                else:
                    tzefi_meuhad.append(self.df.iloc[r - 2][u'צפי מאוחד'])
        else:
            for r in range(2, rows + 2):
                tzefi_meuhad.append(0)
        self.df[u'צפי מאוחד'] = tzefi_meuhad

    def set_3_last_notes(self):
        rows = len(self.df.index)
        for row in range(0, rows):
            last_3_notes = u""
            notes = self.df[u'הערות מגיליון הגביה'][row]
            patern = re.compile(r'\d{2}[/-]\d{2}[/-]\d{2,4}')
            list_of_notes = patern.split(notes)
            list_of_dates = re.findall(r'\d{2}[/-]\d{2}[/-]\d{2,4}', notes)
            list_of_notes = filter(lambda x: x != u'', list_of_notes[-3:])
            list_of_dates = filter(lambda x: x != u'', list_of_dates[-3:])
            for note, date in zip(list_of_notes, list_of_dates):
                last_3_notes = last_3_notes + date + note
            self.df[u'הערות מגיליון הגביה'][row] = last_3_notes

    def arrange_col_order(self):
        self.df = self.df[[u'צוות', u'שם לקוח', u'אמצעי תשלום', u'סוג לקוח', u'לקוח משפטי', u'תשלום ייעוץ חודשי',
                           u'תשלום על בונוס', u'תשלומים מיוחדים', u'מספר התשלומים מעבר לאשראי',
                           u'סכום התשלומים מעבר לאשראי', u'הערות', u'תאריך נקודת הביקורת לבונוס', u'סכום הבונוס',
                           u'צפי לחודש', u'תשלומים נוספים בחודש', u'צפי מאוחד', u'הערות מגיליון הגביה']]

    def change_type_of_col_to_int(self, col_name):
        self.df[col_name] = self.df[col_name].apply(lambda x: int(x))

    def change_type_of_sums_col_to_int(self):
        self.change_type_of_col_to_int(u'תשלום ייעוץ חודשי')
        self.change_type_of_col_to_int(u'תשלום על בונוס')
        self.change_type_of_col_to_int(u'תשלומים מיוחדים')
        self.change_type_of_col_to_int(u'סכום התשלומים מעבר לאשראי')

    def add_mid_sums(self):
        rows = len(self.df.index)
        team = self.df[u'צוות'][0]
        first_index = 0
        last_index = 0
        for r in range(0, rows):
            if r == rows - 1:
                last_index = len(self.df.index) - 1
                self.calc_mid_sums(first_index, last_index, team)
            if self.df[u'צוות'][r + 1] == team:
                last_index = last_index + 1

            else:
                if self.df[u'צוות'][r + 1] != team:
                    new_team = self.df[u'צוות'][r + 1]
                    self.calc_mid_sums(first_index, last_index, team)
                    last_index = last_index + 1
                    first_index = last_index + 1
                    team = new_team
        self.df.sort_index()

    def calc_mid_sums(self, first_index, last_index, team):
        index1 = first_index + 2
        index2 = last_index + 2
        sum_of_month_consultation = '=SUM(F{first_index}:F{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_pay_for_bonus = '=SUM(G{first_index}:G{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_special_payments = '=SUM(H{first_index}:H{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_month_over_credit = '=SUM(J{first_index}:J{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_bonus = '=SUM(M{first_index}:M{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_month_zefi = '=SUM(N{first_index}:N{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_extra_payments_for_month = '=SUM(O{first_index}:O{last_index})'.format(first_index=index1,
                                                                                      last_index=index2)
        sum_of_zefi_meuhad = '=SUM(P{first_index}:P{last_index})'.format(first_index=index1, last_index=index2)
        new_sum_line = pd.DataFrame({u'צוות': [team], u'שם לקוח': [u'סיכום לצוות {team}'.format(team=team)],
                                     u'אמצעי תשלום': [""], u'סוג לקוח': [""], u'לקוח משפטי': [""],
                                     u'תשלום ייעוץ חודשי': [sum_of_month_consultation],
                                     u'תשלום על בונוס': [sum_of_pay_for_bonus],
                                     u'תשלומים מיוחדים': [sum_of_special_payments],
                                     u'מספר התשלומים מעבר לאשראי': [""],
                                     u'סכום התשלומים מעבר לאשראי': [sum_of_month_over_credit],
                                     u'הערות': [""], u'תאריך נקודת הביקורת לבונוס': [""],
                                     u'סכום הבונוס': [sum_of_bonus],
                                     u'צפי לחודש': [sum_of_month_zefi],
                                     u'תשלומים נוספים בחודש': [sum_of_extra_payments_for_month],
                                     u'צפי מאוחד': [sum_of_zefi_meuhad], u'הערות מגיליון הגביה': [""]})
        new_df = concat([self.df.ix[:last_index], new_sum_line, self.df.ix[last_index + 1:]]).reset_index(drop=True)
        new_df.sort_index()
        self.df = new_df
        self.rows_of_sum.append(last_index + 1)

    def add_total_row(self):
        sum_of_month_consultation = '='
        sum_of_pay_for_bonus = '='
        sum_of_special_payments = '='
        sum_of_month_over_credit = '='
        sum_of_bonus = "="
        sum_of_month_zefi = '='
        sum_of_extra_payments_for_month = '='
        sum_of_zefi_meuhad = '='
        for row in range(0, len(self.rows_of_sum) - 1):
            sum_of_month_consultation += "F" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_pay_for_bonus += "G" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_special_payments += "H" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_month_over_credit += "J" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_bonus += "M" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_month_zefi += "N" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_extra_payments_for_month += "O" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_zefi_meuhad += "P" + str(self.rows_of_sum[row] + 2) + "+"
        sum_of_month_consultation = sum_of_month_consultation[:-1]
        sum_of_pay_for_bonus = sum_of_pay_for_bonus[:-1]
        sum_of_special_payments = sum_of_special_payments[:-1]
        sum_of_month_over_credit = sum_of_month_over_credit[:-1]
        sum_of_bonus = sum_of_bonus[:-1]
        sum_of_month_zefi = sum_of_month_zefi[:-1]
        sum_of_extra_payments_for_month = sum_of_extra_payments_for_month[:-1]
        sum_of_zefi_meuhad = sum_of_zefi_meuhad[:-1]
        new_sum_line = pd.DataFrame({u'צוות': [""], u'שם לקוח': [u'סיכום לכל הצוותים:'],
                                     u'אמצעי תשלום': [""], u'סוג לקוח': [""], u'לקוח משפטי': [""],
                                     u'תשלום ייעוץ חודשי': [sum_of_month_consultation],
                                     u'תשלום על בונוס': [sum_of_pay_for_bonus],
                                     u'תשלומים מיוחדים': [sum_of_special_payments],
                                     u'מספר התשלומים מעבר לאשראי': [""],
                                     u'סכום התשלומים מעבר לאשראי': [sum_of_month_over_credit],
                                     u'הערות': [""], u'תאריך נקודת הביקורת לבונוס': [""],
                                     u'סכום הבונוס': [sum_of_bonus], u'צפי לחודש': [sum_of_month_zefi],
                                     u'תשלומים נוספים בחודש': [sum_of_extra_payments_for_month],
                                     u'צפי מאוחד': [sum_of_zefi_meuhad], u'הערות מגיליון הגביה': [""]})
        new_df = concat([self.df, new_sum_line]).reset_index(drop=True)
        self.df = new_df

    def get_df_for_bonus(self):
        query = '''SELECT
        NameCustomer,
        bonusFirstDAteTest AS [Bonus Start Date],
        ISNULL(bonusDAteTest, dateadd(MONTH, Amonth, CONVERT(DATE, DateNew))) AS [Bonus End Date],
        ISNULL(bonusTest/100.0,0) as [Bonus Test],
        DateP,
        DateInvoice,
        NumPay,
        Pay,
        SumS,
        bonus
    FROM
      dbo.tblCustomers
    INNER JOIN (
        SELECT *
        FROM
            (
              SELECT
                IDagreement,
                KodCustomer,
                DateNew,
                bonusFirstDAteTest,
                bonusDAteTest,
                Amonth,
                bonusTest,
                agreementFinish,
                bonus,
                row_number() OVER(PARTITION BY KodCustomer ORDER BY DateNew DESC) AS orderAgreementByDateDesc
                  FROM
                    dbo.tblAgreementConditionAdvice
            ) AS temp
        WHERE temp.orderAgreementByDateDesc = 1
        ) AS tblLastAgreements ON tblLastAgreements.KodCustomer = dbo.tblCustomers.KodCustomer
    INNER JOIN dbo.tblAgreementConditionAdvicePay
          ON tblLastAgreements.IDagreement = dbo.tblAgreementConditionAdvicePay.NumR
    WHERE CustomerStatus = 2
    AND agreementFinish = 0
    ORDER BY NameCustomer'''
        q = self.manager.db_service.search(query=query)
        df = pd.DataFrame.from_records(q)
        # df.to_excel("yaki.xlsx")
        return df

    def get_dict_for_bonus(self):
        df = self.get_df_for_bonus()
        customer = df['NameCustomer'][0]
        last_agreement_bonus_begin_date = df['Bonus Start Date'][0]
        last_agreement_bonus_end_date = df['Bonus End Date'][0]
        bonus_percent = df['Bonus Test'][0]
        sum_pay = 0
        sum_efficiencies = 0
        dict_for_bonus = {}
        dict_for_dates = {}
        for row in range(0, len(df)):
            date = datetime.datetime(
                year=df['Bonus End Date'][row].year, month=df['Bonus End Date'][row].month,
                day=df['Bonus End Date'][row].day)
            if df['Pay'][row] and df['NumPay'][row] is not None and df['NumPay'][row] != 0 and \
                                    last_agreement_bonus_begin_date <= df['DateP'][
                                row] <= last_agreement_bonus_end_date:
                sum_pay += df['Pay'][row]
            if df['SumS'][row] is not None and not pd.isnull(df['DateInvoice'][row]) and \
                                    last_agreement_bonus_begin_date <= datetime.datetime(
                                year=df['DateInvoice'][row].year, month=df['DateInvoice'][row].month,
                                day=1) <= last_agreement_bonus_end_date:
                sum_efficiencies += df['SumS'][row]
            if len(df) - 1 == row:
                bonus_test = sum_efficiencies - (sum_pay * bonus_percent)
                if bonus_test > 0 and df['bonus'][row]:
                    if '%' in df['bonus'][row]:
                        dict_for_bonus[customer] = (bonus_test * float(df['bonus'][row][:-1]) / 100.0) / (1 + self.vat)
                        dict_for_dates[customer] = date
                    else:
                        dict_for_bonus[customer] = (bonus_test * float(df['bonus'][row]) / 100.0) / (1 + self.vat)
                        dict_for_dates[customer] = date
                elif bonus_test < 0 and df['bonus'][row]:
                    dict_for_bonus[customer] = 0
                    dict_for_dates[customer] = date
            else:
                if df['NameCustomer'][row + 1] != customer:
                    bonus_test = sum_efficiencies - (sum_pay * bonus_percent)
                    if bonus_test > 0 and df['bonus'][row]:
                        if '%' in df['bonus'][row]:
                            dict_for_bonus[customer] = (bonus_test * float(df['bonus'][row][:-1]) / 100.0) / (
                                1 + self.vat)
                            dict_for_dates[customer] = date
                        else:
                            dict_for_bonus[customer] = (bonus_test * float(df['bonus'][row]) / 100.0) / (1 + self.vat)
                            dict_for_dates[customer] = date
                    elif bonus_test < 0 and df['bonus'][row]:
                        dict_for_bonus[customer] = 0
                        dict_for_dates[customer] = date
                    customer = df['NameCustomer'][row + 1]
                    last_agreement_bonus_begin_date = df['Bonus Start Date'][row + 1]
                    last_agreement_bonus_end_date = df['Bonus End Date'][row + 1]
                    bonus_percent = df['Bonus Test'][row + 1]
                    sum_pay = 0
                    sum_efficiencies = 0
        return [dict_for_bonus, dict_for_dates]

    def add_col_sum_of_bonus(self):
        dict_for_bonus = self.get_dict_for_bonus()[0]
        self.df[u'סכום הבונוס'] = np.nan
        for index in range(0, len(self.df)):
            if self.df.iloc[index][u'שם לקוח'] in dict_for_bonus.keys():
                self.df.set_value(index, u'סכום הבונוס',
                                  dict_for_bonus.get(self.df.iloc[index][u'שם לקוח']))

    def add_col_date_of_bonus(self):
        dict_for_dates = self.get_dict_for_bonus()[1]
        self.df[u'תאריך נקודת הביקורת לבונוס'] = np.nan
        for index in range(0, len(self.df)):
            if self.df.iloc[index][u'שם לקוח'] in dict_for_dates.keys():
                self.df.set_value(index, u'תאריך נקודת הביקורת לבונוס',
                                  dict_for_dates.get(self.df.iloc[index][u'שם לקוח']))

    def change_types_of_cells(self):
        wb = load_workbook(u'תכנון הצפי לחודש {month}-{year}.xlsx'.format(month=self.chosen_date.month, year=self.chosen_date.year))
        ws = wb.active
        for i in range(2, len(self.df) + 2):
            ws['F{i}'.format(i=i)].number_format = '#,###'
            ws['G{i}'.format(i=i)].number_format = '#,###'
            ws['H{i}'.format(i=i)].number_format = '#,###'
            ws['J{i}'.format(i=i)].number_format = '#,###'
            ws['M{i}'.format(i=i)].number_format = '#,###'
            ws['N{i}'.format(i=i)].number_format = '#,###'
            ws['O{i}'.format(i=i)].number_format = '#,###'
            ws['P{i}'.format(i=i)].number_format = '#,###'
            ws['Q{i}'.format(i=i)].alignment = Alignment(vertical="justify")
        for column in ws.columns:
            for cell in column:
                if not cell.protection.locked:
                    cell.protection = Protection(locked=True)

        file_name = u'תכנון הצפי לחודש {month}-{year}.xlsx'.format(month=self.chosen_date.month, year=self.chosen_date.year)
        wb.save(file_name)

    def set_gvia_month_report_style(self):
        sf = StyleFrame(self.df)
        dict = {u'צוות': 6.38, u'שם לקוח': 7, u'אמצעי תשלום': 6.63, u'סוג לקוח': 4.88,
                u'לקוח משפטי': 6.38, u'תשלום ייעוץ חודשי': 9.63, u'תשלום על בונוס': 10.38,
                u'תשלומים מיוחדים': 7.63, u'מספר התשלומים מעבר לאשראי': 9.5,
                u'סכום התשלומים מעבר לאשראי': 9.63, u'הערות': 6.63,
                u'תאריך נקודת הביקורת לבונוס': 8.63,
                u'סכום הבונוס': 10.38, u'צפי לחודש': 10.25, u'תשלומים נוספים בחודש': 9.5,
                u'צפי מאוחד': 9.38,
                u'הערות מגיליון הגביה': 20.5}
        dict = {key: value + 0.62 for key, value in dict.iteritems()}
        sf.set_column_width_dict(col_width_dict=dict)
        sf.set_row_height(range(1, len(self.df)), 71.25)
        for row in self.rows_of_sum:
            sf.apply_style_by_indexes(indexes_to_style=[sf.index[row]], bg_color=colors.yellow, bold=True)
        return sf

    def export_excel_file(self, sf):
        file_name = u'תכנון הצפי לחודש {month}-{year}.xlsx'.format(month=self.chosen_date.month, year=self.chosen_date.year)
        writer = StyleFrame.ExcelWriter(file_name)
        sf.to_excel(excel_writer=writer, sheet_name=u'דוח גביה חודשי- לקוחות פעילים', right_to_left=True,
                    row_to_add_filters=0,
                    columns_and_rows_to_freeze='A2')
        query_for_gvia_month_non_active_report = '''SELECT
            NameTeam AS [צוות],
            NameCustomer AS [שם לקוח],
            MiddlePay AS [אמצעי תשלום],
            AgreementKind AS [סוג לקוח],
            CASE WHEN proceFinished=0 THEN 'כן' ELSE '' END AS [לקוח משפטי],
            CASE WHEN Conditions='דולר' THEN ISNULL(PayDollar, 0)*4 ELSE ISNULL(PayDollar, 0) END AS [תשלום ייעוץ חודשי],
            ISNULL(SumBonus,0) as [תשלום על בונוס],
            ISNULL(SumSpecial,0) as [תשלומים מיוחדים],
            ISNULL(NumOfDebt,0) as [מספר התשלומים מעבר לאשראי],
            ISNULL(Debt,0) as [סכום התשלומים מעבר לאשראי],
            ISNULL(Remark,'') as [הערות],
            ISNULL(Text, '') AS [הערות מגיליון הגביה]
          FROM
            dbo.tblCustomers
            LEFT JOIN
            dbo.tblTeamList
              ON dbo.tblCustomers.KodTeamCare = dbo.tblTeamList.KodTeamCare
            LEFT JOIN
            (SELECT * FROM
              (
                SELECT
                  IDagreement,
                  KodAgreementKind,
                  KodCustomer,
                  PayProcess,
                  PayDollar,
                  agreementFinish,
                  row_number() OVER(PARTITION BY KodCustomer ORDER BY DateNew DESC) AS orderAgreementByDateDesc
                FROM
                  dbo.tblAgreementConditionAdvice
              ) AS temp
                WHERE temp.orderAgreementByDateDesc = 1) AS tblLastAgreements
              ON dbo.tblCustomers.KodCustomer = tblLastAgreements.KodCustomer
            LEFT JOIN tblJudicialProc
              ON tblCustomers.KodCustomer = tblJudicialProc.kodCustomer
            LEFT JOIN dbo.tblMiddlePay
              ON tblLastAgreements.PayProcess = dbo.tblMiddlePay.KodMiddlePay
            LEFT JOIN dbo.tblAgreementKind
              ON dbo.tblAgreementKind.KodAgreementKind = tblLastAgreements.KodAgreementKind
            LEFT JOIN (
              SELECT DISTINCT NumR, Conditions,
                SUM(CASE WHEN NumPay = 0 AND SumP-Pay>0 AND PayRemark LIKE '%בונוס%' THEN SumP-pay ELSE 0 END) OVER(PARTITION BY NumR) AS SumBonus,
                SUM(CASE WHEN NumPay = 0 AND SumP-Pay>0  AND NOT PayRemark LIKE '%בונוס%' THEN SumP-pay ELSE 0 END) OVER(PARTITION BY NumR) AS SumSpecial,
                SUM(CASE WHEN NumPay > 0 AND SumP-Pay>0  AND DatePay IS NULL AND (CAST('{year}-{month}-{day} 00:00:00' AS DATETIME) > (CASE
               WHEN tblAgreementConditionAdvice.shotef = 1 THEN
                 DateAdd(DAY, tblAgreementConditionAdvice.ashray+1,
                         DateAdd(DAY,-1, Cast(CONVERT(VARCHAR(10),DATEPART(MM, DateAdd(month,1,tblAgreementConditionAdvicePay.DateP)),103)
                                              + '/' + '01/' + CONVERT(VARCHAR(10),DATEPART(YYYY, DateAdd(month,1,tblAgreementConditionAdvicePay.DateP)),103) AS DATETIME)))
               ELSE
                 DateAdd(DAY, tblAgreementConditionAdvice.ashray+1,tblAgreementConditionAdvicePay.DateP)
               END)) THEN 1 ELSE 0 END) OVER(PARTITION BY NumR) AS NumOfDebt,
                SUM(CASE WHEN NumPay > 0 AND SumP-Pay>0  AND DatePay IS NULL AND (CAST('{year}-{month}-{day} 00:00:00' AS DATETIME) > (CASE
               WHEN tblAgreementConditionAdvice.shotef = 1 THEN
                 DateAdd(DAY, tblAgreementConditionAdvice.ashray+1,
                         DateAdd(DAY,-1, Cast(CONVERT(VARCHAR(10),DATEPART(MM, DateAdd(month,1,tblAgreementConditionAdvicePay.DateP)),103)
                                              + '/' + '01/' + CONVERT(VARCHAR(10),DATEPART(YYYY, DateAdd(month,1,tblAgreementConditionAdvicePay.DateP)),103) AS DATETIME)))
               ELSE
                 DateAdd(DAY, tblAgreementConditionAdvice.ashray+1,tblAgreementConditionAdvicePay.DateP)
               END) ) THEN SumP ELSE 0 END) OVER(PARTITION BY NumR) AS Debt,
                STUFF((SELECT ', ' + tblRemarks.PayRemark AS [text()]FROM dbo.tblAgreementConditionAdvicePay tblRemarks WHERE tblRemarks.NumR = dbo.tblAgreementConditionAdvicePay.NumR AND tblRemarks.NumPay = 0 AND SumP-Pay>0 FOR XML PATH('')), 1, 1, '' ) AS Remark
              FROM dbo.tblAgreementConditionAdvicePay
              LEFT JOIN dbo.tblAgreementConditionAdvice
                ON dbo.tblAgreementConditionAdvice.IDagreement = dbo.tblAgreementConditionAdvicePay.NumR


               )AS tblSumAllPay
              ON tblLastAgreements.IDagreement = tblSumAllPay.NumR
            LEFT JOIN (
              SELECT KodCustomer, Text
              FROM tbltaxes
              ) tbltaxes
              ON
                dbo.tblCustomers.KodCustomer = tbltaxes.KodCustomer
          WHERE CustomerStatus != 2
          AND ((MONTH(CAST('{year}-{month}-{day} 00:00:00' AS DATETIME)) = 1 AND YEAR(closeCustomerDate) = YEAR(CAST('{year}-{month}-{day} 00:00:00' AS DATETIME)) - 1 AND MONTH(closeCustomerDate) = 12)
              OR (MONTH(CAST('{year}-{month}-{day} 00:00:00' AS DATETIME)) != 1 AND YEAR(closeCustomerDate) = YEAR(CAST('{year}-{month}-{day} 00:00:00' AS DATETIME)) AND MONTH(closeCustomerDate) = MONTH(CAST('{year}-{month}-{day} 00:00:00' AS DATETIME)) - 1))
          ORDER BY NameTeam'''.format(year=self.year, day=self.day, month=self.month)
        gvia_non_active = GviaMonthNonActiveReport(query_for_gvia_month_non_active_report)
        sf = gvia_non_active.create_non_active_cutomers_gvia_month_report()
        sf.to_excel(excel_writer=writer, sheet_name=u'גביה חודשי- לקוחות לא פעילים', right_to_left=True,
                    row_to_add_filters=0,
                    columns_and_rows_to_freeze='A2')
        writer.save()

    def run_gvia_monthly_report(self):
        self.subtract_vat()
        self.subtract_vat()
        self.add_col_tzefi_for_month(0)
        self.add_col_additional_payments_for_month(0)
        self.add_col_tzefi_meuhad_for_month(0)
        self.add_col_sum_of_bonus()
        self.add_col_date_of_bonus()
        self.set_3_last_notes()
        self.change_type_of_sums_col_to_int()
        self.change_type_of_col_to_int(u'מספר התשלומים מעבר לאשראי')
        self.add_mid_sums()
        self.rows_of_sum.append(len(self.df))
        self.add_total_row()
        self.add_col_tzefi_for_month(1)
        self.add_col_additional_payments_for_month(1)
        self.add_col_tzefi_meuhad_for_month(1)
        self.arrange_col_order()
        sf = self.set_gvia_month_report_style()
        self.export_excel_file(sf)
        self.change_types_of_cells()


if __name__ == '__main__':
    gvia_manager = GviaMonthReport(datetime.datetime.now())
    gvia_manager.subtract_vat()
    gvia_manager.add_col_tzefi_for_month(0)
    gvia_manager.add_col_additional_payments_for_month(0)
    gvia_manager.add_col_tzefi_meuhad_for_month(0)
    gvia_manager.add_col_sum_of_bonus()
    gvia_manager.add_col_date_of_bonus()
    gvia_manager.set_3_last_notes()
    gvia_manager.change_type_of_sums_col_to_int()
    gvia_manager.change_type_of_col_to_int(u'מספר התשלומים מעבר לאשראי')
    gvia_manager.add_mid_sums()
    gvia_manager.rows_of_sum.append(len(gvia_manager.df))
    gvia_manager.add_total_row()
    gvia_manager.add_col_tzefi_for_month(1)
    gvia_manager.add_col_additional_payments_for_month(1)
    gvia_manager.add_col_tzefi_meuhad_for_month(1)
    gvia_manager.arrange_col_order()
    sf = gvia_manager.set_gvia_month_report_style()
    gvia_manager.export_excel_file(sf)
    gvia_manager.change_types_of_cells()
