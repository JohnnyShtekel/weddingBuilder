# coding=utf-8

import pandas as pd
from esg_services import EsgServiceManager

import numpy as np

#
#
# df = pd.DataFrame({
#     'A': ['A1', 'A2', 'A3', 'A4', 'A5', 'A6'],
#     'B': ['B1', 'B2', 'B3', 'B4', 'B5', 'B6'],
#     'C': ['C1', 'C2', 'C3', 'C4', 'C5', 'C6'],
#     'D': ['D1', 'D2', 'D3', 'D4', 'D5', 'D6'],
#     'E': ['E1', 'E2', 'E3', 'E4', 'E5', 'E6'],
#     'F': ['F1', 'F2', 'F3', 'F4', 'F5', 'F6'],
#     'G': ['G1', 'G2', 'G3', 'G4', 'G5', 'G6'],
#     'H': ['H1', 'H2', 'H3', 'H4', 'H5', 'H6'],
# })
#
# df['new'] = np.nan
# df.apply((lambda x: 2), axis=1)
#
# print df
# df1 = df[2:4]
# print df1
#
# s = u'עדי'
# print u'שם: {name}'.format(name=s)

# df = pd.read_excel('Gvia day report new.xlsx', convert_float=True)
# df = df[[u'צוות']]
# df = df.drop_duplicates(u'צוות')
# list = df[u'צוות'].tolist()
# print list

# class tests(object):
#     def __init__(self):
#         self.manager = EsgServiceManager()
#
# tst = tests()
# query = '''SELECT PaySuccessFU, DateFU, dbo.tblCustomers.KodCustomer, NameCustomer,NameTeam
#   FROM dbo.tbltaxes
#   INNER JOIN dbo.tblCustomers
#     ON dbo.tbltaxes.KodCustomer = dbo.tblCustomers.KodCustomer
#   INNER JOIN dbo.tblTeamList
#     ON dbo.tblCustomers.KodTeamCare = dbo.tblTeamList.KodTeamCare
#     WHERE CustomerStatus = 2 AND YEAR(DateFU) = YEAR(getdate()) AND MONTH(DateFU) = MONTH(getdate())
#      AND DAY(DateFU) <= DAY(getdate())
#         ORDER BY NameTeam'''
#
#
# q = tst.manager.db_service.search(query=query)
# df = pd.DataFrame.from_records(q)
# df.to_excel("gvia megvia.xlsx")

# manager = EsgServiceManager()
# columns_list = manager.db_service.search(query="""
# SELECT column_name
#     FROM information_schema.columns
#     WHERE table_name = 'tblMonthlyTarget'
#         AND column_name LIKE 'sumOfTarget%'""")
# columns = map(lambda x: x['column_name'], columns_list)
#
# targets = manager.db_service.search(query='''
# SELECT {targets}
# FROM dbo.tblMonthlyTarget
# WHERE year = YEAR(getdate()) AND month = month(getdate())
# ''')

from openpyxl import load_workbook
from openpyxl.styles import Protection
wb = load_workbook('total gvia.xlsx')
ws = wb.active
for i in range(2,11):
    ws['J{i}'.format(i=i)].number_format = '0.00%'



for column in ws.columns:
    for cell in column:
        if not cell.protection.locked:
            cell.protection = Protection(locked=True)
wb.save('total gvia.xlsx')

