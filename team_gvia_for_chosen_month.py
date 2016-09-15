# coding=utf-8
import pandas as pd
from esg_services import EsgServiceManager
import numpy as np
from StyleFrame import StyleFrame
from pandas import concat
from openpyxl import load_workbook
from openpyxl.styles import Protection


class TeamGviaForMonth(object):
    def __init__(self, requested_year, requested_month):
        self.manager = EsgServiceManager()
        self.df = pd.DataFrame()
        self.rows_of_sum = []
        self.df_from_sql_for_gvia_megvia_and_tzefi = np.nan
        query_for_vat = "SELECT * from dbo.TaxRate"
        q = self.manager.db_service.search(query=query_for_vat)
        df_for_vat = pd.DataFrame.from_records(q)
        self.vat = df_for_vat['Tax'][len(df_for_vat) - 1]
        self.month = requested_month
        self.year = requested_year

    def add_col_team(self):
        query = '''SELECT NameTeam [צוות]
            FROM dbo.tblTeamList
            WHERE NameTeam  NOT LIKE 'כללי'
            ORDER BY NameTeam'''
        q = self.manager.db_service.search(query=query)
        self.df = pd.DataFrame.from_records(q)

    def get_paid_payment_for_consultation_and_gvia_from_excel_daily_report(self):
        df = pd.read_excel(u' דוח גביה יומי {month}-{year}.xlsx'.format(month=self.month, year=self.year),
                           convert_float=True, sheetname=u'דוח גביה יומי חיובי')
        df = df[[u'צוות', u'שם לקוח', u'סוג לקוח', u'תשלום ששולם עד היום לייעוץ', u'תשלום ששולם עד היום לגביה']]
        list_of_sum_rows = []
        for row in range(0, len(df)):
            if df[u'שם לקוח'][row].startswith(u'סיכום'):
                list_of_sum_rows.append(row)
        df.drop(df.index[list_of_sum_rows], inplace=True)
        df.index = range(len(df))
        return df

    def add_gvia_cols_for_3_kinds_of_customers(self):
        df = self.get_paid_payment_for_consultation_and_gvia_from_excel_daily_report()
        list_for_regular_customer_col = []
        list_for_consultation_customer_col = []
        list_for_project_customer_col = []
        team = df[u'צוות'][0]
        regular_sum = 0
        consultation_sum = 0
        project_sum = 0
        for row in range(0, len(df)):
            if df[u'צוות'][row] == team:
                if str(df[u'תשלום ששולם עד היום לייעוץ'][row]) == "nan":
                    df[u'תשלום ששולם עד היום לייעוץ'][row] = 0.0
                if df[u'סוג לקוח'][row] == u'ES - בסיס הצלחה':
                    regular_sum += df[u'תשלום ששולם עד היום לייעוץ'][row]
                elif df[u'סוג לקוח'][row] == u'ES-יעוץ':
                    consultation_sum += df[u'תשלום ששולם עד היום לייעוץ'][row]
                elif df[u'סוג לקוח'][row] == u'פרוייקטלי':
                    project_sum += df[u'תשלום ששולם עד היום לייעוץ'][row]
            if row == len(df) - 1:
                list_for_regular_customer_col.append(regular_sum)
                list_for_consultation_customer_col.append(consultation_sum)
                list_for_project_customer_col.append(project_sum)
            elif df[u'צוות'][row + 1] != team:
                team = df[u'צוות'][row + 1]
                list_for_regular_customer_col.append(regular_sum)
                list_for_consultation_customer_col.append(consultation_sum)
                list_for_project_customer_col.append(project_sum)
                regular_sum = 0
                consultation_sum = 0
                project_sum = 0
        self.df[u'גביה מרגיל'] = list_for_regular_customer_col
        self.df[u'גביה מייעוץ'] = list_for_consultation_customer_col
        self.df[u'גביה מפרויקטלי'] = list_for_project_customer_col

    def add_col_total_gvia(self):
        list_of_formulas = []
        for row in range(0, len(self.df)):
            formula = "=B{row}+C{row}+D{row}".format(row=row + 2)
            list_of_formulas.append(formula)
        self.df[u'גביה כוללת'] = list_of_formulas

    def add_col_target(self, requestd_year, requested_month):
        list_of_targets = []
        list_of_indexes_for_teams_targets = [3, 2, 1, 10, 6, 4, 5, 7, 8]
        # TODO: If you want to get last month report, change the month below:
        query = '''SELECT *
        FROM dbo.tblMonthlyTarget
        WHERE year = {requestd_year} AND month = {requested_month}'''.format(requestd_year=requestd_year, requested_month=requested_month)
        q = self.manager.db_service.search(query=query)
        df = pd.DataFrame.from_records(q)
        for index in list_of_indexes_for_teams_targets:
            list_of_targets.append(df['sumOfTarget{index}'.format(index=index)][0])
        self.df[u'יעד'] = list_of_targets

    def add_col_percentage_of_completion(self):
        self.df[u'אחוז ביצוע'] = np.nan
        for row in range(0, len(self.df)):
            self.df[u'אחוז ביצוע'][row] = "=(IF(F{row}>0,E{row}/F{row},0))".format(row=row + 2)

    def add_row_for_yesum(self):
        formula = "=SUM(({col}2:{col}5),({col}8,{col}10))"
        new_yesum_row = pd.DataFrame({u'צוות': [u'סה"כ יישום'], u'גביה מרגיל': [formula.format(col='B')],
                          u'גביה מייעוץ': [formula.format(col='C')], u'גביה מפרויקטלי': [formula.format(col='D')],
                          u'גביה כוללת': [formula.format(col='E')],
                          u'יעד': [formula.format(col='F')], u'אחוז ביצוע': ["=(IF(F11>0,E11/F11,0))"]})
        self.df = concat([self.df, new_yesum_row]).reset_index(drop=True)

    def add_row_for_gvia(self):
        df = self.get_paid_payment_for_consultation_and_gvia_from_excel_daily_report()
        regular_sum = 0
        consultation_sum = 0
        project_sum = 0
        for row in range(0, len(df)):
            if str(df[u'תשלום ששולם עד היום לגביה'][row]) == "nan":
                df[u'תשלום ששולם עד היום לגביה'][row] = 0.0
            if df[u'סוג לקוח'][row] == u'ES - בסיס הצלחה':
                regular_sum += df[u'תשלום ששולם עד היום לגביה'][row]
            elif df[u'סוג לקוח'][row] == u'ES-יעוץ':
                consultation_sum += df[u'תשלום ששולם עד היום לגביה'][row]
            elif df[u'סוג לקוח'][row] == u'פרוייקטלי':
                project_sum += df[u'תשלום ששולם עד היום לגביה'][row]
        formula = "=SUM({col}2:{col}10)"
        new_gvia_row = pd.DataFrame({u'צוות': [u'סה"כ גביה'], u'גביה מרגיל': [regular_sum],
                                      u'גביה מייעוץ': [consultation_sum],
                                      u'גביה מפרויקטלי': [project_sum],
                                      u'גביה כוללת': ["=B12+C12+D12"],
                                      u'יעד': [formula.format(col='F')], u'אחוז ביצוע': ["=(IF(F12>0,E12/F12,0))"]})
        self.df = concat([self.df, new_gvia_row]).reset_index(drop=True)



    def order_columns(self):
        self.df = self.df[[u'צוות', u'גביה מרגיל', u'גביה מייעוץ', u'גביה מפרויקטלי', u'גביה כוללת', u'יעד', u'אחוז ביצוע']]

    def change_types_of_cells(self):
        wb = load_workbook(u'דוח גביה יומי מחלקתי {month}-{year}.xlsx'.format(month=self.month, year=self.year))
        ws = wb.active
        for i in range(2, 13):
            ws['G{i}'.format(i=i)].number_format = '0%'
            ws['B{i}'.format(i=i)].number_format = '#,###'
            ws['C{i}'.format(i=i)].number_format = '#,###'
            ws['D{i}'.format(i=i)].number_format = '#,###'
            ws['E{i}'.format(i=i)].number_format = '#,###'
            ws['F{i}'.format(i=i)].number_format = '#,###'
        for column in ws.columns:
            for cell in column:
                if not cell.protection.locked:
                    cell.protection = Protection(locked=True)
        wb.save(u'דוח גביה יומי מחלקתי {month}-{year}.xlsx'.format(month=self.month, year=self.year))

    def create_report_file(self):
        file_name = u'דוח גביה יומי מחלקתי {month}-{year}.xlsx'.format(month=self.month,year=self.year)
        sf = StyleFrame(self.df)
        writer = StyleFrame.ExcelWriter(file_name)
        sf.to_excel(excel_writer=writer, sheet_name=u'דוח גביה יומי לפי צוותים', right_to_left=True,
                    row_to_add_filters=0,
                    columns_and_rows_to_freeze='A2')
        writer.save()
        return file_name

    def run_report(self):
        self.add_col_team()
        self.add_gvia_cols_for_3_kinds_of_customers()
        self.add_col_total_gvia()
        self.add_col_target(2016, 8)
        self.add_col_percentage_of_completion()
        self.add_row_for_yesum()
        self.add_row_for_gvia()
        self.order_columns()
        file_name = self.create_report_file()
        self.change_types_of_cells()
        return file_name


if __name__ == '__main__':

    gvia_teams = TeamGviaForMonth(2016, 8)
    gvia_teams.add_col_team()
    gvia_teams.add_gvia_cols_for_3_kinds_of_customers()
    gvia_teams.add_col_total_gvia()
    gvia_teams.add_col_target(2016, 8)
    gvia_teams.add_col_percentage_of_completion()
    gvia_teams.add_row_for_yesum()
    gvia_teams.add_row_for_gvia()
    gvia_teams.order_columns()

    sf = StyleFrame(gvia_teams.df)
    writer = StyleFrame.ExcelWriter(u'דוח גביה יומי מחלקתי {month}-{year}.xlsx'.format(month=gvia_teams.month,
                                                                                       year=gvia_teams.year))
    sf.to_excel(excel_writer=writer, sheet_name=u'דוח גביה יומי לפי צוותים', right_to_left=True, row_to_add_filters=0,
            columns_and_rows_to_freeze='A2')
    writer.save()
    gvia_teams.change_types_of_cells()








