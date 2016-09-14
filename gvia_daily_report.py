# coding=utf-8
import pandas as pd
from esg_services import EsgServiceManager
from StyleFrame import StyleFrame, colors
import datetime
import re
import numpy as np
from pandas import concat
from openpyxl import load_workbook
from openpyxl.styles import Protection


class GviaDaily(object):
    def __init__(self, status, chosen_date, for_previous_month):
        self.chosen_date = chosen_date
        self.for_previous_month = for_previous_month
        self.year = str(chosen_date.year)
        self.month = '0' + str(chosen_date.month) if chosen_date.month < 10 else str(chosen_date.month)
        self.day = '0' + str(chosen_date.day) if chosen_date.day < 10 else str(chosen_date.day)

        # TODO: If you want to get report for different month(not the current) change MONTH(getdate()) in the query below
        query = '''SELECT
          ISNULL(NameTeam, taxTeam) AS [צוות],
          dbo.tblCustomers.NameCustomer AS [שם לקוח],
          MiddlePay AS [אמצעי תשלום],
          AgreementKind AS [סוג לקוח],
          DateFU AS [תאריך לביצוע],
          CASE WHEN proceFinished=0 THEN 'כן' ELSE '' END AS [לקוח משפטי],
          CASE WHEN Conditions='דולר' THEN ISNULL(PayDollar, 0)*4 ELSE ISNULL(PayDollar, 0) END AS [תשלום ייעוץ חודשי],
          ISNULL(Text, '') AS [הערות מגיליון הגביה]
        FROM
          dbo.tblCustomers
          LEFT JOIN
          dbo.tblTeamList
            ON dbo.tblCustomers.KodTeamCare = dbo.tblTeamList.KodTeamCare
          LEFT JOIN
          (SELECT * FROM
            (
              SELECT
                IDagreement,
                KodAgreementKind,
                KodCustomer,
                PayProcess,
                PayDollar,
                agreementFinish,
                row_number() OVER(PARTITION BY KodCustomer ORDER BY DateNew DESC) AS orderAgreementByDateDesc
              FROM
                dbo.tblAgreementConditionAdvice
            ) AS temp
              WHERE temp.orderAgreementByDateDesc = 1) AS tblLastAgreements
            ON dbo.tblCustomers.KodCustomer = tblLastAgreements.KodCustomer
          LEFT JOIN tblJudicialProc
            ON tblCustomers.KodCustomer = tblJudicialProc.kodCustomer
          LEFT JOIN dbo.tblMiddlePay
            ON tblLastAgreements.PayProcess = dbo.tblMiddlePay.KodMiddlePay
          LEFT JOIN dbo.tblAgreementKind
            ON dbo.tblAgreementKind.KodAgreementKind = tblLastAgreements.KodAgreementKind
          LEFT JOIN (
            SELECT DISTINCT NumR, Conditions
            FROM dbo.tblAgreementConditionAdvicePay
            LEFT JOIN dbo.tblAgreementConditionAdvice
              ON dbo.tblAgreementConditionAdvice.IDagreement = dbo.tblAgreementConditionAdvicePay.NumR
             )AS tblSumAllPay
            ON tblLastAgreements.IDagreement = tblSumAllPay.NumR
          LEFT JOIN (
            SELECT KodCustomer, Text, DateFU
            FROM tbltaxes
            ) tbltaxes
            ON
              dbo.tblCustomers.KodCustomer = tbltaxes.KodCustomer
          LEFT JOIN (
                      SELECT DISTINCT
                        NameTeam AS [taxTeam],
                        KodCustomer,
                        1 AS [taxsThisMonth]
                      FROM dbo.tbltaxesPay
                        LEFT JOIN dbo.tblTeamList
                        ON dbo.tbltaxesPay.KodTeamCare = dbo.tblTeamList.KodTeamCare
                      WHERE YEAR(DatePay) = YEAR(CAST('{year}-{month}-{day} 00:00:00' AS DATETIME)) AND MONTH(DatePay) = MONTH(CAST('{year}-{month}-{day} 00:00:00' AS DATETIME))
                    ) taxesPay
            ON taxesPay.KodCustomer = dbo.tblCustomers.KodCustomer
        WHERE ((CustomerStatus = 2 OR (CustomerStatus != 2  AND taxesPay.taxsThisMonth = 1)))
        AND ((agreementFinish = 0 OR (CustomerStatus != 2  AND taxesPay.taxsThisMonth = 1)))
        ORDER BY [צוות]'''.format(day=self.day, month=self.month, year=self.year)
        self.manager = EsgServiceManager()
        print query
        # self.day = datetime.datetime.today().day
        # self.month = datetime.datetime.today().month
        # self.year = datetime.datetime.today().year
        # TODO: If you want to get report for different month(not the current) change file name below to the right file!
        month_report_file_name = u'תכנון הצפי לחודש {month}-{year}.xlsx'.format(month=self.chosen_date.month,
                                                                                year=self.chosen_date.year)
        self.df_from_month_report = pd.read_excel(month_report_file_name, convert_float=True,
                                                  sheetname=u'דוח גביה חודשי- לקוחות פעילים')
        # self.df_from_month_report = pd.read_excel(u'תכנון הצפי לחודש אוגוסט 2016.xlsx', convert_float=True, sheetname=u'צפי לחודש אוגוסט 2016')
        if status == 'report_for_hani':
            self.df = pd.DataFrame.from_records(self.manager.db_service.search(query=query))
        else:
            self.df = np.nan
        self.df_positive = np.nan
        self.df_negative = np.nan
        self.rows_of_sum = []
        self.list_of_positive_customers = []

    def remove_sum_rows(self):
        list_of_sum_rows = []
        for row in range(0, len(self.df_from_month_report)):
            if self.df_from_month_report[u'שם לקוח'][row].startswith(u'סיכום'):
                list_of_sum_rows.append(row)
        self.df_from_month_report.drop(self.df_from_month_report.index[list_of_sum_rows], inplace=True)
        self.df_from_month_report.index = range(len(self.df_from_month_report))

    def round_col_payment_for_month_consultation(self):
        print self.df
        self.df[u'תשלום ייעוץ חודשי'] = self.df[u'תשלום ייעוץ חודשי'].apply(lambda x: round(x))

    def calc_month_tzefi(self):
        month_tzefi_list = []
        for row in range(0, len(self.df_from_month_report)):
            if self.df_from_month_report[u'לקוח משפטי'][row] == u'כן':
                month_tzefi_list.append(0)
            else:
                if self.df_from_month_report[u'מספר התשלומים מעבר לאשראי'][row] > 3:
                    month_tzefi_list.append(0)
                else:
                    month_tzefi_list.append(self.df_from_month_report[u'תשלום ייעוץ חודשי'][row])
        return month_tzefi_list

    def calc_month_extra_payments(self):
        month_extra_payments_list = []
        for row in range(0, len(self.df_from_month_report)):
            if self.df[u'לקוח משפטי'][row] == u'כן':
                month_extra_payments_list.append(0)
            else:
                month_extra_payments_list.append(self.df_from_month_report[u'תשלום על בונוס'][row] +
                                                 self.df_from_month_report[u'תשלומים מיוחדים'][row] +
                                                 self.df_from_month_report[u'סכום התשלומים מעבר לאשראי'][row])
        return month_extra_payments_list

    def add_col_month_tzefi_from_original_month_report(self):
        month_tzefi_list = []
        for tzefi, extra in zip(self.calc_month_tzefi(), self.calc_month_extra_payments()):
            month_tzefi_list.append(round(tzefi + extra))
        self.df_from_month_report[u'צפי לחודש'] = month_tzefi_list
        dict_for_tzefi = self.df_from_month_report.set_index(u'שם לקוח')[u'צפי לחודש'].to_dict()
        self.df[u'צפי לחודש'] = np.nan
        for index in range(0, len(self.df)):
            if self.df.iloc[index][u'שם לקוח'] in dict_for_tzefi.keys():
                self.df.set_value(index, u'צפי לחודש',
                                  dict_for_tzefi.get(self.df.iloc[index][u'שם לקוח']))

    def add_col_month_tzefi(self):
        self.df[u'צפי לחודש'] = np.nan
        if str(self.df_from_month_report[u'צפי לחודש'][0]) == "nan":
            self.add_col_month_tzefi_from_original_month_report()
        else:
            month_tzefi_list = []
            for row in range(0, len(self.df_from_month_report)):
                month_tzefi_list.append(round(self.df_from_month_report[u'צפי מאוחד'][row]))
            self.df_from_month_report[u'צפי לחודש'] = month_tzefi_list
            dict_for_tzefi = self.df_from_month_report.set_index(u'שם לקוח')[u'צפי לחודש'].to_dict()
            self.df[u'צפי לחודש'] = np.nan
            for index in range(0, len(self.df)):
                if self.df.iloc[index][u'שם לקוח'] in dict_for_tzefi.keys():
                    self.df.set_value(index, u'צפי לחודש',
                                      dict_for_tzefi.get(self.df.iloc[index][u'שם לקוח']))

    def get_df_of_fees(self):
        dates = []
        query = '''SELECT NameCustomer, feeTeam, DatePay, feeTax
                FROM dbo.tblCustomers LEFT JOIN dbo.tbltaxesPay
                  ON dbo.tbltaxesPay.KodCustomer = dbo.tblCustomers.KodCustomer ORDER BY NameCustomer'''
        q = self.manager.db_service.search(query=query)
        df = pd.DataFrame.from_records(q)
        # today = datetime.datetime.today()
        for index, row in df.iterrows():
            # TODO: If you want to get report for different month(not the current) change today.month below
            if row['DatePay'] is not None and row['DatePay'].year == self.chosen_date.year and row[
                'DatePay'].month == self.chosen_date.month:
                date = row['DatePay']
                dates.append(date)
            else:
                date = datetime.datetime(1920, 1, 1, 00, 00, 00)
                dates.append(date)
        df['date_pay'] = dates
        df.drop(['DatePay'], axis=1, inplace=True)
        old_date_for_compare = datetime.datetime(1920, 1, 1, 00, 00, 00)
        df = df[df.date_pay > old_date_for_compare]
        df.drop(['date_pay'], axis=1, inplace=True)
        return df

    def add_cols_for_fees(self):
        df = self.get_df_of_fees()
        self.add_col_paid_payment_for_consultation(df)
        self.add_col_paid_payment_for_gvia(df)

    def get_dict_for_fee_team(self, df):
        df = df.reset_index()
        del df['index']
        customer = df['NameCustomer'][0]
        for row in range(0, len(df)):
            if df['feeTeam'][row] > 0 and df['NameCustomer'][row] not in self.list_of_positive_customers:
                self.list_of_positive_customers.append(df['NameCustomer'][row])
            if row + 1 != len(df) and df['NameCustomer'][row + 1] != customer:
                customer = df['NameCustomer'][row + 1]
        df_grouped = pd.DataFrame(df.groupby('NameCustomer').feeTeam.sum()).reset_index()
        dict_for_feeTeam = df_grouped.set_index('NameCustomer')['feeTeam'].to_dict()
        return dict_for_feeTeam

    def add_col_paid_payment_for_consultation(self, df):
        dict_of_feeTeam = self.get_dict_for_fee_team(df)
        self.df[u'תשלום ששולם עד היום לייעוץ'] = np.nan
        for index in range(0, len(self.df)):
            if self.df.iloc[index][u'שם לקוח'] in dict_of_feeTeam.keys():
                self.df.set_value(index, u'תשלום ששולם עד היום לייעוץ',
                                  round(dict_of_feeTeam.get
                                        (self.df.iloc[index][u'שם לקוח'])))
        dict_for_consultation = self.df.set_index(u'שם לקוח')[u'תשלום ששולם עד היום לייעוץ'].to_dict()
        self.df[u'תשלום ששולם עד היום לייעוץ'] = np.nan
        for index in range(0, len(self.df)):
            if self.df.iloc[index][u'שם לקוח'] in dict_for_consultation.keys():
                self.df.set_value(index, u'תשלום ששולם עד היום לייעוץ',
                                  dict_for_consultation.get(self.df.iloc[index][u'שם לקוח']))

    def get_dict_for_fee_tax(self, df):
        df_grouped = pd.DataFrame(df.groupby('NameCustomer').feeTax.sum()).reset_index()
        dict_for_feeTax = df_grouped.set_index('NameCustomer')['feeTax'].to_dict()
        return dict_for_feeTax

    def add_col_paid_payment_for_gvia(self, df):
        dict_of_feeTax = self.get_dict_for_fee_tax(df)
        self.df[u'תשלום ששולם עד היום לגביה'] = np.nan
        for index in range(0, len(self.df)):
            if self.df.iloc[index][u'שם לקוח'] in dict_of_feeTax.keys():
                self.df.set_value(index, u'תשלום ששולם עד היום לגביה', round(dict_of_feeTax.get(self.
                                                                                                df.iloc[index][
                                                                                                    u'שם לקוח'])))
        dict_for_gvia = self.df.set_index(u'שם לקוח')[u'תשלום ששולם עד היום לגביה'].to_dict()
        self.df[u'תשלום ששולם עד היום לגביה'] = np.nan
        for index in range(0, len(self.df)):
            if self.df.iloc[index][u'שם לקוח'] in dict_for_gvia.keys():
                self.df.set_value(index, u'תשלום ששולם עד היום לגביה',
                                  dict_for_gvia.get(self.df.iloc[index][u'שם לקוח']))

    def add_col_execution_date(self):
        dates = []
        for index, row in self.df.iterrows():
            if row[u'תאריך לביצוע'] is not None:
                date = row[u'תאריך לביצוע']
                date = date.date()
                dates.append(date)
            else:
                date = np.nan
                dates.append(date)
        self.df[u'תאריך לביצוע'] = dates

    def add_cols_for_hani(self):
        self.df[u'תשובות לחני'] = np.nan
        self.df[u'הערות חני'] = np.nan

    def add_col_notes_from_gvia_sheet(self):
        notes_for_column = []
        rows = len(self.df)
        for row in range(0, rows):
            last_3_notes = u""
            notes = self.df[u'הערות מגיליון הגביה'][row]
            patern = re.compile(r'\d{2}[/-]\d{2}[/-]\d{2,4}')
            list_of_notes = patern.split(notes)
            list_of_dates = re.findall(r'\d{2}[/-]\d{2}[/-]\d{2,4}', notes)
            list_of_notes = filter(lambda x: x != u'', list_of_notes[-3:])
            list_of_dates = filter(lambda x: x != u'', list_of_dates[-3:])
            for note, date in zip(list_of_notes, list_of_dates):
                last_3_notes = last_3_notes + date + note
            notes_for_column.append(last_3_notes)
        self.df[u'הערות מגיליון הגביה'] = notes_for_column

    def add_col_notes_from_gvia_month(self):
        self.df[u'הערות משורת החיוב בסטטוס'] = np.nan
        rows = len(self.df)
        for row in range(0, rows):
            for row1 in range(0, len(self.df_from_month_report)):
                if self.df[u'שם לקוח'][row] == self.df_from_month_report[u'שם לקוח'][row1]:
                    self.df[u'הערות משורת החיוב בסטטוס'][row] = self.df_from_month_report[u'הערות'][row1]

    def calc_col_tzefi_left(self, df):
        list_for_col = []
        rows = len(df.index)
        for index in range(0, rows):
            if index not in self.rows_of_sum:
                list_for_col.append('=IF(F{row}-G{row}<0,0,F{row}-G{row})'.format(row=index + 2))
            else:
                list_for_col.append(df.iloc[index][u'צפי שנותר'])
        df[u'צפי שנותר'] = list_for_col
        return df

    def add_empty_col_tzefi_left(self):
        list_for_col = []
        rows = len(self.df.index)
        for index in range(0, rows):
            list_for_col.append(0)
        self.df[u'צפי שנותר'] = list_for_col

    def order_columns(self, df):
        df = df[[u'צוות', u'שם לקוח', u'סוג לקוח', u'לקוח משפטי', u'תשלום ייעוץ חודשי', u'צפי לחודש',
                 u'תשלום ששולם עד היום לייעוץ', u'צפי שנותר', u'תשלום ששולם עד היום לגביה', u'הערות מגיליון הגביה',
                 u'הערות משורת החיוב בסטטוס', u'תאריך לביצוע', u'אמצעי תשלום', u'תשובות לחני', u'הערות חני']]
        return df

    def add_mid_sums(self, df):
        rows = len(df.index)
        team = df.iloc[0][u'צוות']
        first_index = 0
        last_index = 0
        for r in range(0, rows):
            if r == rows - 1:
                last_index = len(df.index) - 1
                df = self.calc_mid_sums(df, first_index, last_index, team)
            elif df[u'צוות'][r + 1] == team:
                last_index += 1
            else:
                if df[u'צוות'][r + 1] != team:
                    new_team = df[u'צוות'][r + 1]
                    df = self.calc_mid_sums(df, first_index, last_index, team)
                    last_index += 1
                    first_index = last_index + 1
                    team = new_team
        df = df.sort_index()
        return df

    def calc_mid_sums(self, df, first_index, last_index, team):
        index1 = first_index + 2
        index2 = last_index + 2
        sum_of_month_consultation = '=SUM(E{first_index}:E{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_month_zefi = '=SUM(F{first_index}:F{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_paid_payment_for_consultation = '=SUM(G{first_index}:G{last_index})'.format(first_index=index1,
                                                                                           last_index=index2)
        sum_of_zefi_left = '=SUM(H{first_index}:H{last_index})'.format(first_index=index1, last_index=index2)
        sum_of_paid_payment_for_gvia = '=SUM(I{first_index}:I{last_index})'.format(first_index=index1,
                                                                                   last_index=index2)
        new_sum_line = pd.DataFrame({u'צוות': [team], u'שם לקוח': [u'סיכום לצוות {team}'.format(team=team)],
                                     u'סוג לקוח': [""], u'לקוח משפטי': [""],
                                     u'תשלום ייעוץ חודשי': [sum_of_month_consultation],
                                     u'צפי לחודש': [sum_of_month_zefi],
                                     u'תשלום ששולם עד היום לייעוץ': [sum_of_paid_payment_for_consultation],
                                     u'צפי שנותר': [sum_of_zefi_left],
                                     u'תשלום ששולם עד היום לגביה': [sum_of_paid_payment_for_gvia],
                                     u'הערות מגיליון הגביה': [""], u'הערות משורת החיוב בסטטוס': [""],
                                     u'תאריך לביצוע': [""],
                                     u'אמצעי תשלום': [""],
                                     u'תשובות לחני': [""],
                                     u'הערות חני': [""]})
        new_df = concat([df.ix[:last_index], new_sum_line, df.ix[last_index + 1:]]).reset_index(drop=True)
        new_df.sort_index()
        self.rows_of_sum.append(last_index + 1)
        return new_df

    def add_total_row(self, df):
        sum_of_month_consultation = '='
        sum_of_month_zefi = '='
        sum_of_paid_payment_for_consultation = '='
        sum_of_zefi_left = '='
        sum_of_paid_payment_for_gvia = '='
        for row in range(0, len(self.rows_of_sum)):
            sum_of_month_consultation += "E" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_month_zefi += "F" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_paid_payment_for_consultation += "G" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_zefi_left += "H" + str(self.rows_of_sum[row] + 2) + "+"
            sum_of_paid_payment_for_gvia += "I" + str(self.rows_of_sum[row] + 2) + "+"
        sum_of_month_consultation = sum_of_month_consultation[:-1]
        sum_of_month_zefi = sum_of_month_zefi[:-1]
        sum_of_paid_payment_for_consultation = sum_of_paid_payment_for_consultation[:-1]
        sum_of_zefi_left = sum_of_zefi_left[:-1]
        sum_of_paid_payment_for_gvia = sum_of_paid_payment_for_gvia[:-1]
        new_sum_line = pd.DataFrame({u'צוות': [""], u'שם לקוח': [u'סיכום לכל הצוותים:'],
                                     u'סוג לקוח': [""], u'לקוח משפטי': [""],
                                     u'תשלום ייעוץ חודשי': [sum_of_month_consultation],
                                     u'צפי לחודש': [sum_of_month_zefi],
                                     u'תשלום ששולם עד היום לייעוץ': [sum_of_paid_payment_for_consultation],
                                     u'צפי שנותר': [sum_of_zefi_left],
                                     u'תשלום ששולם עד היום לגביה': [sum_of_paid_payment_for_gvia],
                                     u'הערות מגיליון הגביה': [""], u'הערות משורת החיוב בסטטוס': [""],
                                     u'תאריך לביצוע': [""],
                                     u'אמצעי תשלום': [""],
                                     u'תשובות לחני': [""],
                                     u'הערות חני': [""]})
        new_df = concat([df, new_sum_line]).reset_index(drop=True)
        return new_df

    def change_rows_for_positive_test(self):
        self.df[u'תשלום ששולם עד היום לייעוץ'][67] *= -1
        self.df[u'תשלום ששולם עד היום לייעוץ'][5] *= -1
        self.df[u'תשלום ששולם עד היום לייעוץ'][9] *= -1

    def create_positive_df(self, status):
        cond_for_negative = self.df[u'תשלום ששולם עד היום לייעוץ'] < 0
        self.df_positive = self.df[~cond_for_negative]
        self.df_positive = self.df_positive.reset_index(drop=True)
        self.df_positive = self.add_mid_sums(self.df_positive)
        if status == 'report_for_hani':
            self.df_positive = self.add_total_row(self.df_positive)
            self.rows_of_sum.append(self.rows_of_sum[len(self.rows_of_sum) - 1] + 1)
        self.df_positive = self.calc_col_tzefi_left(self.df_positive)
        self.df_positive = self.order_columns(self.df_positive)

    def create_negative_df(self, status):
        self.rows_of_sum = []
        cond_for_negative = self.df[u'תשלום ששולם עד היום לייעוץ'] < 0
        self.df_negative = self.df[cond_for_negative]
        self.df_negative = self.df_negative.reset_index(drop=True)
        if len(self.df_negative.index) != 0:
            self.df_negative = self.add_mid_sums(self.df_negative)
            if status == 'report_for_hani':
                self.df_negative = self.add_total_row(self.df_negative)
                self.rows_of_sum.append(self.rows_of_sum[len(self.rows_of_sum) - 1] + 1)
            self.df_negative = self.calc_col_tzefi_left(self.df_negative)
            self.df_negative = self.order_columns(self.df_negative)

    def create_df(self):
        self.round_col_payment_for_month_consultation()
        self.remove_sum_rows()
        self.add_empty_col_tzefi_left()
        self.add_cols_for_fees()
        self.add_col_execution_date()
        self.add_cols_for_hani()
        self.add_col_notes_from_gvia_sheet()
        self.add_col_month_tzefi()
        self.add_col_notes_from_gvia_month()
        self.df = self.order_columns(self.df)

    def apply_style_on_sum_rows(self, sf):
        for row in self.rows_of_sum:
            sf.apply_style_by_indexes(indexes_to_style=[sf.index[row]], bg_color=colors.yellow, bold=True)

    def create_separated_df(self, file_name, status):
        self.create_positive_df(status)
        sf_positive = StyleFrame(self.df_positive)
        self.apply_style_on_sum_rows(sf_positive)
        writer = StyleFrame.ExcelWriter(file_name)
        sf_positive.to_excel(excel_writer=writer, sheet_name=u'דוח גביה יומי חיובי', right_to_left=True,
                             row_to_add_filters=0, columns_and_rows_to_freeze='A2')
        self.create_negative_df(status)
        if len(self.df_negative.index) != 0:
            sf_negative = StyleFrame(self.df_negative)
            self.apply_style_on_sum_rows(sf_negative)
            sf_negative.to_excel(excel_writer=writer, sheet_name=u'דוח גביה יומי שלילי', right_to_left=True,
                                 row_to_add_filters=0, columns_and_rows_to_freeze='A2')
        writer.save()

    def create_report_for_each_team(self):
        gvia_for_teams = GviaDaily('report for teams', self.chosen_date, self.for_previous_month)
        rows = len(self.df.index)
        team = self.df.iloc[0][u'צוות']
        first_index = 0
        last_index = 0
        for r in range(0, rows):
            if r == rows - 1:
                last_index = r
                gvia_for_teams.df = self.df[first_index:last_index + 1]
                gvia_for_teams.df.reset_index(drop=True)
                file_name = u' דוח גביה יומי {day}-{month}-{year} צוות {team}.xlsx'.format(team=team,
                                                                                           day=self.chosen_date.day,
                                                                                           month=self.chosen_date.month,
                                                                                           year=self.chosen_date.year)
                gvia_for_teams.create_separated_df(file_name, 'report_for_teams')
            elif self.df[u'צוות'][r + 1] == team:
                last_index += 1
            else:
                if self.df[u'צוות'][r + 1] != team:
                    gvia_for_teams = GviaDaily('report for teams', self.chosen_date, self.for_previous_month)
                    gvia_for_teams.df = self.df[first_index:last_index + 1]
                    gvia_for_teams.df.reset_index(drop=True)
                    file_name = u' דוח גביה יומי {day}-{month}-{year} צוות {team}.xlsx'.format(team=team,
                                                                                               day=self.chosen_date.day,
                                                                                               month=self.chosen_date.month,
                                                                                               year=self.chosen_date.year)
                    gvia_for_teams.create_separated_df(file_name, 'report_for_teams')
                    new_team = self.df[u'צוות'][r + 1]
                    last_index += 1
                    first_index = last_index
                    team = new_team

    def change_types_of_cells(self):

        wb = load_workbook(
            u' דוח גביה יומי {day}-{month}-{year}.xlsx'.format(day=self.chosen_date.day, month=self.chosen_date.month,
                                                               year=self.chosen_date.year))
        ws = wb.active
        for i in range(2, len(self.df_positive) + 2):
            ws['E{i}'.format(i=i)].number_format = '#,###'
            ws['F{i}'.format(i=i)].number_format = '#,###'
            ws['G{i}'.format(i=i)].number_format = '#,###'
            ws['H{i}'.format(i=i)].number_format = '#,###'
            ws['I{i}'.format(i=i)].number_format = '#,###'
        for column in ws.columns:
            for cell in column:
                if not cell.protection.locked:
                    cell.protection = Protection(locked=True)
        wb.save(
            u' דוח גביה יומי {day}-{month}-{year}.xlsx'.format(day=self.chosen_date.day, month=self.chosen_date.month,
                                                               year=self.chosen_date.year))

    def send_mail_to_managers(self):
        files = []
        emails = ['YakiN@esg.co.il', 'HaniP@esg.co.il', 'rond@esg.co.il']
        file_name = u' דוח גביה יומי {day}-{month}-{year}.xlsx'.format(day=self.chosen_date.day,
                                                                       month=self.chosen_date.month,
                                                                       year=self.chosen_date.year)
        files.append(file_name)
        # TODO: change mail_to to emails list!!!
        mail_to = ['yakin@esg.co.il', 'HaniP@esg.co.il', 'ortalb@esg.co.il']
        mail_from = ('ESG Server', 'server@esg.co.il')
        subject = u'דוח גביה יומי {day}-{month}-{year}'.format(day=self.chosen_date.day, month=self.chosen_date.month,
                                                               year=self.chosen_date.year)
        response = self.manager.mail_service.send_mail(mail_from=mail_from, mail_to=mail_to, files=files,
                                                       subject=subject, text=u'מצורף דוח גביה יומי')
        print response.content

    def run_daily_report(self):
        self.create_df()
        self.create_separated_df(u' דוח גביה יומי {day}-{month}-{year}.xlsx'.format(day=self.chosen_date.day,
                                                                                    month=self.chosen_date.month,
                                                                                    year=self.chosen_date.year),
                                 'report_for_hani')
        self.create_report_for_each_team()
        self.change_types_of_cells()
        if not self.for_previous_month:
            self.send_mail_to_managers()


if __name__ == '__main__':
    current_date = datetime.datetime(2016,8,25)
    gvia_daily = GviaDaily('report_for_hani', current_date, False)
    gvia_daily.create_df()
    # gvia_daily.change_rows_for_positive_test()
    gvia_daily.create_separated_df(u' דוח גביה יומי {day}-{month}-{year}.xlsx'.format(day=current_date.day,
                                                                                      month=current_date.month,
                                                                                      year=current_date.year),
                                   'report_for_hani')
    gvia_daily.create_report_for_each_team()
    gvia_daily.change_types_of_cells()
    # gvia_daily.send_mail_to_managers()
